# -*- coding: utf-8 -*-
import pandas as pd
import numpy as np
import logging
import base64
import io
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from collections import defaultdict
from datetime import datetime, date, timedelta

from odoo import api, fields, models, _, Command
from odoo.tools import DEFAULT_SERVER_DATETIME_FORMAT
from odoo.exceptions import UserError, ValidationError, AccessError, RedirectWarning

from odoo.addons.l10n_co_exogenous_information_reporting.tools.utils import _column_name_field
from odoo.addons.l10n_co_exogenous_information_reporting.tools.utils import dian_countries_codes
from odoo.addons.l10n_co_exogenous_information_reporting.tools.utils import dian_department_codes
from odoo.addons.l10n_co_exogenous_information_reporting.tools.utils import _check_dv


_logger = logging.getLogger(__name__)


class L10ncoExogenousFormatSetting(models.Model):
    _name = "l10n_co.exogenous_format_setting"
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _description = "Exogenous formats configuration model"
    _rec_name = 'format_id'

    @api.constrains('is_it_with_date_range', 'date_start', 'date_end')
    def _check_dates(self):
        for record in self:
            if record.is_it_with_date_range and record.date_start and record.date_end:
                if record.date_start > record.date_end:
                    raise ValidationError(
                        _("The start date must be less than the end date"))

                if record.date_start.year != record.date_end.year:
                    raise ValidationError(
                        _("The start date and end date must be in the same year"))

    active = fields.Boolean(string='Active', default=True, tracking=True)
    company_id = fields.Many2one(
        comodel_name='res.company', string='Company', default=lambda self: self.env.company.id)
    format_id = fields.Many2one(
        comodel_name='l10n_co.exogenous_format', string='Format', tracking=True)
    format_setting_line_ids = fields.One2many(
        comodel_name='l10n_co.exogenous_format_setting_line', inverse_name='format_setting_id', string='Format Setting Lines')
    apply_concepts = fields.Boolean(
        string='Apply Concepts', related='format_id.apply_concepts', readonly=True, store=True)

    is_it_with_date_range = fields.Boolean(
        string='Is it with date range?', related='format_id.is_it_with_date_range', readonly=True, store=True)
    date_start = fields.Date(string='Start Date')
    date_end = fields.Date(string='End Date')

    journal_ids = fields.Many2many(
        comodel_name='account.journal', string='Journals to be excluded')
    partner_ids = fields.Many2many(
        comodel_name='res.partner', string='Contacts to be excluded')

    binary_file = fields.Binary(string='File', tracking=False)
    binary_file_name = fields.Char(string='File name', tracking=False)
    
    def _get_fields_many2one(self, fields_contact):
        """
            Función que nos permite obtener de los campo de Odoo de contacto cuales son de tipo Many2one
        """
        return fields_contact.mapped('field_odoo_id').filtered(lambda field: field.ttype == 'many2one')

    def _get_last_day_pass_year(self):
        """
            Función que nos permite obtener el ultimo día del año pasado
        """
        return date(date.today().year, 1, 1) - timedelta(days=1)

    def _dynamic_search_read(self, model, domain, fields, order=None):
        """
            Esta función nos permite hacer search_read de manera dinamica
        """
        result = self.env[model].search_read(
            domain=domain,
            fields=fields,
            order=order,
        )
        return result

    def _get_type_documents_by_format(self):
        """
            Esta función nos permite obtener de acuerdo al formato los tipos de documentos de la dian
        """
        model = 'l10n_co.exogenous_document_type'
        domain = [('document_type_table_ids', 'in',
                   self.format_id.document_type_table_id.id), ('type_document_id', '!=', False)]
        fields = ['type_document_id', 'code']
        return self._dynamic_search_read(model=model, domain=domain, fields=fields)

    def _get_columns_account_move_line_by_format(self):
        fields_use_account_move_line = self._get_fields_use_account_move_line()
        if self.format_id.code in ('1007', '1008'):
            return fields_use_account_move_line + ['partner_id', 'account_id']
        return fields_use_account_move_line + ['date', 'partner_id', 'account_id','move_id']

    def _get_accounts_by_format(self):
        vals = {}
        for setting_line in self.format_setting_line_ids:
            if self.format_id.apply_concepts:
                for account in setting_line.concept_id.field_account_ids.account_ids:
                    if account.id not in vals:
                        vals[account.id] = setting_line.concept_id.code
            else:
                for account in setting_line.format_field_id.field_account_ids.account_ids:
                    if account.id not in vals:
                        vals[account.id] = setting_line.format_field_id.id

        return vals

    def _normalice_data_dataframe_partner(self, fields_many2one, df_partners, df_type_documents, df_other_info, fields_to_clear_contact, fields_to_clear_company):
        for key, value in fields_many2one.items():
            if key in df_partners.columns and key in df_other_info.columns:
                mapping_dict = {d['id']: d.get(
                    value, None) for element in df_other_info[key] for d in element}
                df_partners[key] = df_partners[key].apply(
                    lambda x: mapping_dict.get(x[0] if isinstance(x, (list, tuple)) else x, x))
        
        mapping_dict = dict(zip(df_type_documents['type_document_id'].apply(
            lambda x: x[1]), df_type_documents['code']))
        

        df_partners['l10n_latam_identification_type_id'] = df_partners['l10n_latam_identification_type_id'].apply(
            lambda x: mapping_dict.get(x[1] if isinstance(x, (list, tuple)) else x, x))

        for index, row in df_partners.iterrows():
            fields_to_clear = fields_to_clear_company if row['is_company'] else fields_to_clear_contact
            for column in fields_to_clear:
                df_partners.at[index, column] = None

        return df_partners

    def _normalice_and_merge_data_dataframe_account_move_line(self, df_account_move_lines, df_partners):
        
        df_account_move_lines['partner_id'] = df_account_move_lines['partner_id'].apply(
            lambda x: x[0])

        # Realizar la mezcla basada en la columna partner_id y la columna id
        df_result = pd.merge(df_account_move_lines, df_partners,
                             left_on='partner_id', right_on='id', how='left')

        # Eliminar la columna id duplicada, si es necesario
        df_result.drop('id_y', axis=1, inplace=True)
        df_result.drop('id_x', axis=1, inplace=True)
        df_result.drop('partner_id', axis=1, inplace=True)
        
        return df_result

    def _get_field_name_second_field_many2one(self, fields_contact):
        vals = dict()
        for field in fields_contact.filtered(lambda field: field.ttype == 'many2one'):
            if field.field_odoo_id.name not in vals:
                vals[field.field_odoo_id.name] = field.field_odoo_internal_id.name
            else:
                vals[field.field_odoo_id.name] = list(
                    vals.get(field.field_odoo_id.name)).append(field.field_odoo_internal_id.name)
        return vals

    def _get_information_by_account_move_line(self, accounts_ids):

        domain = [('parent_state', '=', 'posted'), ('company_id', '=',
                                                    self.company_id.id), ('account_id', 'in', accounts_ids)]

        fields = self._get_columns_account_move_line_by_format()
        order = 'date asc'

        if not self.is_it_with_date_range:
            domain.append(('date', '<=', self._get_last_day_pass_year()))
        else:
            domain += [('date', '>=', self.date_start),
                       ('date', '<=', self.date_end)]

        if self.journal_ids:
            domain.append(('journal_id', 'not in', self.journal_ids.ids))

        if self.partner_ids:
            domain += [('partner_id', 'not in', self.partner_ids.ids), ('partner_id', '!=', False)]
        else:
            domain.append(('partner_id', '!=', False))

        partner_ids = set()
        account_move_lines = self._dynamic_search_read(
            model='account.move.line', domain=domain, fields=fields, order=order)

        for aml in account_move_lines:
            if aml['partner_id']:
                partner_ids.add(aml['partner_id'][0])

        return partner_ids, account_move_lines

    def _get_values_form_many2one(self, data_source, fields_many2one):
        """ 
            Funcion que nos permite de acuerdo a una informacion obtenida por un search_read
            y con base a unos campos que son de tipo many2one obtener los ids de esos modelos
            y llevarlos en una lista
        """
        values_fields_many2one = dict()
        for ds in data_source:
            for field_many2one in fields_many2one:
                if not isinstance(ds[field_many2one], bool):
                    if field_many2one not in values_fields_many2one:
                        values_fields_many2one[field_many2one] = [
                            ds[field_many2one][0]]
                    else:
                        values_fields_many2one[field_many2one].append(
                            ds[field_many2one][0])
        return values_fields_many2one

    def _get_values_form_dynamic_search_read(self, search_read_by_dynamic_model, field_odoo_internal):
        values = list()
        for value in search_read_by_dynamic_model:
            values.append(value[field_odoo_internal.name])
        return values

    def _get_data_from_dynamic_model(self, partner_info, fields_contact):

        fields_many2one_odoo = self._get_fields_many2one(fields_contact)
        fields_names_many2one_odoo = fields_many2one_odoo.mapped("name")
        values_fields_many2one = self._get_values_form_many2one(
            partner_info, fields_names_many2one_odoo)

        list_fields_name_and_models = fields_contact.mapped('field_odoo_id').filtered(
            lambda field: field.ttype == 'many2one').mapped(lambda field: {field.name: field.relation})
        values_by_dynamic_models = dict()

        for dict_field_name_and_model in list_fields_name_and_models:
            for values in dict_field_name_and_model:
                if values_fields_many2one.get(values, False):
                    field_odoo_internal = fields_contact.filtered(
                        lambda field_line: field_line.field_odoo_id.name == values).mapped('field_odoo_internal_id')

                    model = dict_field_name_and_model[values]
                    domain = [
                        ('id', 'in', list(set(values_fields_many2one.get(values))))]
                    fields = field_odoo_internal.mapped('name')
                    search_read_by_dynamic_model = self._dynamic_search_read(
                        model=model, domain=domain, fields=fields)

                    if values not in values_by_dynamic_models:
                        values_by_dynamic_models[values] = search_read_by_dynamic_model
                    else:
                        values_by_dynamic_models[values].extend(
                            search_read_by_dynamic_model)

        return values_by_dynamic_models

    def _get_information_partner(self, partner_ids, fields_contact):

        if not fields_contact:
            return self._show_message_error(f"No fields are configured for the company: {self.company_id.name}")

        fields_odoo = fields_contact.mapped('field_odoo_id.name')
        fields_odoo.append('is_company')

        model = 'res.partner'
        domain = [('id', 'in', list(partner_ids))]
        fields = fields_odoo

        partner_info = self._dynamic_search_read(
            model=model, domain=domain, fields=fields)

        other_info_contact = self._get_data_from_dynamic_model(
            partner_info, fields_contact)
        return partner_info, other_info_contact

    def _get_accounts(self, format_setting_line):
        if self.format_id.apply_concepts:
            return format_setting_line.mapped('concept_id').mapped('field_account_ids').mapped('account_ids')
        return format_setting_line.mapped('format_field_id').mapped('field_account_ids').mapped('account_ids')

    def _get_field_concept(self):
        return self.env.ref('l10n_co_exogenous_information_reporting.l10n_co_exogenous_format_field_cpt', raise_if_not_found=True)

    def _get_columns_ordered(self, format_fields) -> list:
        vals = list()
        for format_field in format_fields:
            if format_field.id == self._get_field_concept().id:
                vals.append('account_id')
            elif format_field.field_odoo_id:
                vals.append(format_field.field_odoo_id.name)
        return vals

    def _get_fields_to_clear_contact(self, fields_contact):
        return list(set(fields_contact.mapped('field_odoo_id').mapped('name')) - set(fields_contact.filtered(lambda field: field.source == 'contact' and field.applies_to_contact).mapped('field_odoo_id').mapped('name')))

    def _get_fields_to_clear_company(self, fields_contact):
        return list(set(fields_contact.mapped('field_odoo_id').mapped('name')) - set(fields_contact.filtered(lambda field: field.source == 'contact' and field.applies_to_company).mapped('field_odoo_id').mapped('name')))

    def find_column_position(self, sheet, column_name):
        for col_idx in range(1, sheet.max_column + 1):
            if sheet.cell(row=1, column=col_idx).value == column_name:
                return col_idx
        return None  # Retorna None si no se encuentra la columna

    def _format_field_by_accumulated(self, format_setting_line):
        vals = dict()
        if self.format_id.apply_concepts:
            if format_setting_line.format_field_id.name not in vals:
                accounts_accumulated_by = dict()
                for field_account in format_setting_line.concept_id.field_account_ids:
                    for account in field_account.account_ids:
                        accounts_accumulated_by[(account.id, account.display_name)] = field_account.name
                vals[format_setting_line.format_field_id.name] = accounts_accumulated_by
        else:
            accounts_accumulated_by = dict()
            for field_account in format_setting_line.format_field_id.field_account_ids:
                for account in field_account.account_ids:
                    accounts_accumulated_by[(account.id, account.display_name)] = field_account.name
            vals[format_setting_line.format_field_id.name] = accounts_accumulated_by
        return vals

    def _get_fields_use_account_move_line(self):
        return ['credit', 'debit', 'balance', 'tax_base_amount']

    def _get_fields_format_id(self, format_setting_line):
        return format_setting_line.mapped('format_field_id').mapped('name')

    def _create_original_dataframe(self, format_fields):
        return pd.DataFrame(columns=format_fields.mapped('name'))

    def _get_fields_odoo_and_format(self, fields_contact):
        vals = dict()
        if self.format_id.apply_concepts:
            vals['account_id'] = self._get_field_concept().name

        for field_contact in fields_contact: 
            vals[field_contact.field_odoo_id.name] = field_contact.name

        return vals

    def _get_fields_fill_smaller_amount(self):
        return {
            self.env.ref('l10n_co_exogenous_information_reporting.l10n_co_exogenous_format_field_tdoc', raise_if_not_found=True).name : '43',
            self.env.ref('l10n_co_exogenous_information_reporting.l10n_co_exogenous_format_field_nid', raise_if_not_found=True).name : '222222222',
            self.env.ref('l10n_co_exogenous_information_reporting.l10n_co_exogenous_format_field_raz', raise_if_not_found=True).name : 'cuantías menores',
        }
    

    def _generate_row_by_smaller_amount(self, row_smaller_amount, df):
        new_row = pd.DataFrame(row_smaller_amount, index=[0])
        new_row_columns = df.columns
        new_row = new_row.reindex(columns=new_row_columns)
        df = pd.concat([df, new_row], ignore_index=True)
        return df


    def _get_names_columns_direction_contact(self):
        return {
            self.env.ref('l10n_co_exogenous_information_reporting.l10n_co_exogenous_format_field_dpto', raise_if_not_found=True).name: dian_department_codes,
            #self.env.ref('l10n_co_exogenous_information_reporting.l10n_co_exogenous_format_field_mun', raise_if_not_found=True).name: dian_cities_by_department,
            self.env.ref('l10n_co_exogenous_information_reporting.l10n_co_exogenous_format_field_pais', raise_if_not_found=True).name: dian_countries_codes
        }

    def generate_and_download_report(self):
        # Check if format settings are available
        if not self.format_setting_line_ids:
            return self._show_message_error(f"No format settings configured for format: {self.format_id.code}")

        # Retrieve format fields
        format_fields = self.env['l10n_co.exogenous_format_field'].search(
            [('format_ids', 'in', self.format_id.id)], order="sequence asc")

        if not format_fields:
            return self._show_message_error(f"No fields configured for format: {self.format_id.code}")

        # Filter fields based on source
        fields_contact = format_fields.filtered(lambda ff: ff.source == 'contact')

        # Get fields to clear for contact and company
        fields_to_clear_contact = self._get_fields_to_clear_contact(fields_contact)
        fields_to_clear_company = self._get_fields_to_clear_company(fields_contact)

        # Get unique keys for format
        unique_keys_by_format = format_fields.filtered(lambda ff: ff.is_unique_key).mapped(
            'field_odoo_id').mapped('name')

        unique_keys_by_format_field = format_fields.filtered(lambda ff: ff.is_unique_key).mapped('name')

        if self.format_id.apply_concepts:
            unique_keys_by_format = unique_keys_by_format + ['account_id']

        wb = Workbook()
        ws = wb.active
        ws = _column_name_field(format_fields.mapped('name'), ws)

        # Create original dataframe
        original_dataframe = self._create_original_dataframe(format_fields)
        columnas_originales = original_dataframe.columns
        row_smaller_amount = self._get_fields_fill_smaller_amount()
        names_columns_direction_contact = self._get_names_columns_direction_contact()
        concepts = self._get_accounts_by_format()

        get_information = list()

        for setting_line in self.format_setting_line_ids:
            accounts_ids = self._get_accounts(setting_line).ids
            partner_ids, account_move_lines = self._get_information_by_account_move_line(accounts_ids)

            _logger.info(setting_line)
            if not account_move_lines or not partner_ids:
                get_information.append(False)
                continue
            else:
                get_information.append(True)

            partners, other_info_contact = self._get_information_partner(partner_ids, fields_contact)

            if isinstance(partners, dict) and partners.get('tag', False):
                return partners

            df_type_documents = pd.DataFrame(self._get_type_documents_by_format())
            df_account_move_lines = pd.DataFrame(account_move_lines)
            df_partners = pd.DataFrame(partners)
            df_partners = df_partners.fillna('')
            df_partners = df_partners.replace(False, '')
            df_other_info = pd.DataFrame([other_info_contact])

            df_partners = self._normalice_data_dataframe_partner(
                self._get_field_name_second_field_many2one(fields_contact),
                df_partners, df_type_documents, df_other_info, fields_to_clear_contact, fields_to_clear_company
            )

            df_account_move_lines = self._normalice_and_merge_data_dataframe_account_move_line(
                df_account_move_lines, df_partners)

            format_field_accumulated = self._format_field_by_accumulated(setting_line)

            def get_column_value(row, key):
                column_to_get = format_field_accumulated.get(key, None)
                if isinstance(column_to_get, dict):
                    # Obtén el account_id de la fila actual
                    account_id = row['account_id']
                    # Obtén el valor correspondiente al account_id
                    operation = column_to_get.get(account_id, None)
                    if operation:
                        # Mapea las operaciones a las columnas correspondientes
                        operation_to_column = {
                            'credit': 'credit',
                            'debit': 'debit',
                            'balance': 'balance',
                            'tax_base_amount': 'tax_base_amount'
                        }
                        # Obtén la columna correspondiente a la operación
                        column = operation_to_column.get(operation, None)
                        if column:
                            # Calcula el valor de la columna
                            return row[column]
                return 0.0

            columns_ordered = self._get_columns_ordered(format_fields)
            columns_ordered = columns_ordered + [setting_line.format_field_id.name]

            clave_diccionario, valor_diccionario = next(iter(format_field_accumulated.items()))

            df_account_move_lines[clave_diccionario] = df_account_move_lines.apply(
                lambda row: get_column_value(row, clave_diccionario), axis=1)
            
            df_account_move_lines['account_id'] = df_account_move_lines['account_id'].apply(lambda x: concepts.get(x[0], x[0]))

            operations = {column: 'first' if column not in self._get_fields_format_id(setting_line) else 'sum'
                        for column in columns_ordered}

            df_account_move_lines_copy = df_account_move_lines.copy(deep=True)
            filtered_grouped = df_account_move_lines_copy.groupby(unique_keys_by_format)

            result = filtered_grouped.agg(operations).round(0)
            result = result.rename(columns=self._get_fields_odoo_and_format(fields_contact))

            if self.format_id.applying_smaller_amounts:
                smaller_amounts = result.loc[
                    result[setting_line.format_field_id.name] < self.format_id.smaller_ammounts,
                    setting_line.format_field_id.name
                ].sum()
                result = result.loc[
                    result[setting_line.format_field_id.name] >= self.format_id.smaller_ammounts]

                
                if self.format_id.apply_concepts and smaller_amounts > 0:
                    row_smaller_amount = self._get_fields_fill_smaller_amount()
                    row_smaller_amount.update({setting_line.format_field_id.name: smaller_amounts})
                    row_smaller_amount.update({self._get_field_concept().name: setting_line.concept_id.code})
                    result = self._generate_row_by_smaller_amount(row_smaller_amount, result)
                elif not self.format_id.apply_concepts and smaller_amounts > 0:
                    row_smaller_amount.update({setting_line.format_field_id.name: smaller_amounts})

            duplicados = original_dataframe.merge(result, on=unique_keys_by_format_field, how='inner')
            filas_duplicadas = result.loc[result.index.intersection(duplicados.index)]
            for index, fila_duplicada in filas_duplicadas.iterrows():
                filtro = (original_dataframe[unique_keys_by_format_field] == fila_duplicada[unique_keys_by_format_field]).all(axis=1)
                original_dataframe.loc[filtro, fila_duplicada.index] = fila_duplicada.values

            filas_nuevas = result.loc[~result.index.isin(result.index.intersection(duplicados.index))]
            original_dataframe = pd.concat([original_dataframe, filas_nuevas], ignore_index=True)


        if not any(get_information):
            return self._show_message_error("No information found with these parameters")

        if not self.format_id.apply_concepts and self.format_id.applying_smaller_amounts:
            original_dataframe = self._generate_row_by_smaller_amount(row_smaller_amount, original_dataframe)

        original_dataframe = original_dataframe.reindex(columns=columnas_originales)

        
        for column in names_columns_direction_contact:
            replacement_dict = {key.lower(): value for key, value in names_columns_direction_contact[column].items()}
            if column in original_dataframe.columns:
                original_dataframe[column] = (original_dataframe[column].astype(str).str.lower().map(replacement_dict))

        def procesar_identificacion(row):
            nit = row[field_identification_number_name]

            # Validar que nit no esté vacío o sea None
            if not nit or (isinstance(nit, str) and len(nit.strip()) == 0):
                return pd.Series([nit, None])

            # Convertir a string y limpiar espacios
            nit = str(nit).strip()

            if '-' in nit:
                numero, dv_actual = nit.split('-', 1)
            elif len(nit) > 1:
                numero, dv_actual = nit[:-1], nit[-1]
            else:
                # Si solo tiene 1 carácter, no hay DV
                return pd.Series([nit, None])

            dv_nuevo = _check_dv(numero) if 'Digito de Verificación' in original_dataframe.columns else None
            return pd.Series([numero, dv_nuevo if dv_nuevo != dv_actual else dv_actual])

        filter_mask = original_dataframe[self.env.ref('l10n_co_exogenous_information_reporting.l10n_co_exogenous_format_field_tdoc', raise_if_not_found=True).name] == '31'
        field_identification_number_name = self.env.ref('l10n_co_exogenous_information_reporting.l10n_co_exogenous_format_field_nid', raise_if_not_found=True).name
        field_dv_name = self.env.ref('l10n_co_exogenous_information_reporting.l10n_co_exogenous_format_field_dv', raise_if_not_found=True).name if 'Digito de Verificación' in original_dataframe.columns else None

        filtered_dataframe = original_dataframe.loc[filter_mask]
        if not filtered_dataframe.empty:
            results = filtered_dataframe.apply(procesar_identificacion, axis=1)
            results.index = filtered_dataframe.index

            if field_dv_name:
                results.columns = [field_identification_number_name, field_dv_name]
                original_dataframe.loc[results.index, [field_identification_number_name, field_dv_name]] = results[
                    [field_identification_number_name, field_dv_name]
                ]
            else:
                original_dataframe.loc[results.index, field_identification_number_name] = results.iloc[:, 0]

        for row_data in dataframe_to_rows(original_dataframe, index=False, header=False):
            ws.append(row_data)

        output = io.BytesIO()
        wb.save(output)
        file_name = f"exogena_{self.format_id.code}_{datetime.now().strftime(DEFAULT_SERVER_DATETIME_FORMAT)}"
        self.binary_file = base64.b64encode(output.getvalue())
        self.binary_file_name = f"{file_name}.xlsx"


    def _show_message_error(self, message):
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'type': 'danger',
                'sticky': True,
                'message': _(message),
            }
        }
