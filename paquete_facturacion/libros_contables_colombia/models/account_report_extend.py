# -*- coding: utf-8 -*-
"""
Extensión de account.report para Libros Contables Colombia.

Filtros personalizados siguiendo el patrón nativo de Odoo 18:
1. Definir campo filter_* en el modelo
2. Crear método _init_options_* para inicializar opciones
3. Exponer en get_report_information() con 'show_*' en filters dict

Filtros implementados:
- filter_hide_partners_no_movement: Ocultar terceros sin movimiento
- filter_account_range: Filtro por rango de cuentas (desde/hasta/excluir)
- filter_withholding_filters: Filtros de retención colombianos
"""

import io
import base64
import logging
from itertools import groupby

from odoo import models, fields, api, _
import markupsafe

try:
    import xlsxwriter
except ImportError:
    xlsxwriter = None

try:
    from openpyxl import load_workbook as openpyxl_load_workbook
except ImportError:
    openpyxl_load_workbook = None

_logger = logging.getLogger(__name__)

# Lista de handlers colombianos del módulo libros_contables_colombia
COLOMBIAN_HANDLER_MODELS = [
    'account.trial.balance.puc.report.handler',
    'account.trial.balance.partner.report.handler',
    'account.auxiliary.book.colombia.handler',
    'account.ledger.book.colombia.handler',
    'account.journal.book.colombia.handler',
    'account.inventory.book.colombia.handler',
    'co.comparative.balance.report.handler',
    'co.comparative.income.report.handler',
    'co.cash.flow.report.handler',
    'co.equity.changes.report.handler',
    'co.financial.indicators.report.handler',
    'account.capital.difference.handler',
    # Handlers de impuestos colombianos
    'account.tax.report.general.handler',
    'account.tax.report.iva.handler',
    'account.tax.report.retefuente.handler',
    'account.tax.report.reteiva.handler',
    'account.tax.report.reteica.handler',
    'account.tax.report.inc.handler',
    'account.tax.report.consumo.handler',
    'account.tax.report.otros.handler',
]


class AccountReportExtend(models.Model):
    _inherit = 'account.report'

    # =========================================================================
    # CAMPOS DE FILTRO (Paso 1 del patrón nativo)
    # Siguiendo el patrón de filter_hide_0_lines en account_reports
    # =========================================================================

    # Filtro para ocultar terceros sin movimiento (tipo Selection como filter_hide_0_lines)
    filter_hide_partners_no_movement = fields.Selection(
        string="Ocultar terceros sin movimiento",
        selection=[
            ('by_default', "Activado por defecto"),
            ('optional', "Opcional"),
            ('never', "Nunca")
        ],
        default='never',  # Por defecto deshabilitado para no afectar otros reportes
        help="Oculta terceros que solo tienen saldo inicial pero no débitos ni créditos en el período."
    )

    # Filtro por rango de cuentas contables (inputs manuales)
    filter_account_range = fields.Boolean(
        string="Filtro por rango de cuentas",
        default=False,
        help="Habilita el filtro por rango de cuentas contables (desde/hasta) con opción de exclusión."
    )

    # Filtro por cuentas contables (selector múltiple como filter_partner)
    filter_account = fields.Boolean(
        string="Filtro por Cuentas",
        compute=lambda x: x._compute_report_option_filter('filter_account'),
        readonly=False,
        store=True,
        depends=['root_report_id', 'section_main_report_ids'],
        help="Habilita el filtro por cuentas contables específicas usando selector múltiple."
    )

    # Filtro por impuestos (selector múltiple como filter_partner)
    filter_tax = fields.Boolean(
        string="Filtro por Impuestos",
        compute=lambda x: x._compute_report_option_filter('filter_tax'),
        readonly=False,
        store=True,
        depends=['root_report_id', 'section_main_report_ids'],
        help="Habilita el filtro por impuestos específicos usando selector múltiple."
    )

    # =========================================================================
    # FILTROS DE RETENCIÓN / IMPUESTOS COLOMBIANOS
    # =========================================================================
    filter_withholding_filters = fields.Boolean(
        string="Filtros de Retención",
        default=False,
        help="Habilita los filtros de tipo de operación (compra/venta), tipo de concepto y conceptos específicos de retención."
    )
    filter_operation_type = fields.Boolean(
        string="Filtro por Tipo de Operación",
        default=False,
        help="Habilita el filtro por tipo de operación (Compras, Ventas, NC Proveedor, NC Cliente)."
    )
    filter_concept_type = fields.Boolean(
        string="Filtro por Tipo de Concepto",
        default=False,
        help="Habilita el filtro por tipo de concepto de retención (ReteFuente, ReteIVA, ReteICA, IVA, INC)."
    )
    filter_withholding_concept = fields.Boolean(
        string="Filtro por Concepto de Retención",
        default=False,
        help="Habilita el filtro por concepto específico de retención desde tax_base_threshold."
    )
    filter_person_type = fields.Boolean(
        string="Filtro por Tipo de Persona",
        default=False,
        help="Habilita el filtro por tipo de persona (Persona Jurídica PJ / Persona Natural PN)."
    )
    filter_product_type = fields.Boolean(
        string="Filtro por Tipo de Producto",
        default=False,
        help="Habilita el filtro por tipo de producto (Servicios / Bienes)."
    )
    filter_city = fields.Boolean(
        string="Filtro por Ciudad",
        default=False,
        help="Habilita el filtro por ciudad del tercero (res.partner.city_id)."
    )

    # =========================================================================
    # FILTRO PARA OCULTAR CUENTAS SIN MOVIMIENTO (Libro Auxiliar)
    # =========================================================================
    filter_hide_accounts_no_movement = fields.Selection(
        string="Ocultar cuentas sin movimiento",
        selection=[
            ('by_default', "Activado por defecto"),
            ('optional', "Opcional"),
            ('never', "Nunca")
        ],
        default='never',  # Por defecto deshabilitado para no afectar otros reportes
        help="Oculta cuentas que solo tienen saldo inicial pero no tienen débitos ni créditos en el período seleccionado."
    )

    # =========================================================================
    # FILTRO PARA JERARQUÍA PUC COLOMBIANO
    # Estructura: Clase (1) > Grupo (2) > Cuenta (4) > Subcuenta (6) > Auxiliar (8+)
    # =========================================================================
    filter_puc_hierarchy = fields.Selection(
        string="Jerarquía PUC",
        selection=[
            ('by_default', "Activado por defecto"),
            ('optional', "Opcional"),
            ('never', "Nunca")
        ],
        default='never',
        help="Agrupa las cuentas según la estructura del PUC colombiano:\n"
             "- Clase: 1 dígito (ej: 1 Activo)\n"
             "- Grupo: 2 dígitos (ej: 11 Disponible)\n"
             "- Cuenta: 4 dígitos (ej: 1105 Caja)\n"
             "- Subcuenta: 6 dígitos (ej: 110505 Caja General)\n"
             "- Auxiliar: 8+ dígitos (ej: 11050501 Caja General ME)"
    )

    # =========================================================================
    # OPCIONES DE FORMATO Y EXPORTACIÓN
    # =========================================================================

    filter_color_formatting = fields.Selection(
        string="Formato de Colores",
        selection=[
            ('by_default', "Activado por defecto"),
            ('optional', "Opcional"),
            ('never', "Nunca")
        ],
        default='optional',
        help="Aplica colores a números: rojo para negativos, verde para positivos."
    )

    filter_secondary_currency = fields.Boolean(
        string="Mostrar Moneda Secundaria",
        default=False,
        help="Habilita la opción de mostrar valores en moneda secundaria."
    )

    filter_initial_balance_date = fields.Boolean(
        string="Fecha de Último Movimiento en Saldo Inicial",
        default=False,
        help="Muestra la fecha del último movimiento en los saldos iniciales."
    )

    filter_export_wizard = fields.Boolean(
        string="Wizard de Exportación Avanzado",
        default=True,
        help="Habilita el wizard de exportación con opciones avanzadas."
    )

    # =========================================================================
    # MÉTODOS _init_options_* (Paso 2 del patrón nativo)
    # Estos métodos son descubiertos automáticamente por get_options()
    # usando reflexión: for attr in dir(self) if attr.startswith('_init_options_')
    # =========================================================================

    def _init_options_hide_partners_no_movement(self, options, previous_options=None):
        """
        Inicializa la opción para ocultar terceros sin movimiento.

        Sigue el patrón exacto de _init_options_hide_0_lines del código nativo:
        1. Verificar si el filtro está habilitado (filter_* != 'never')
        2. Usar valor previo si existe, sino usar default del campo
        3. Si filter_* == 'never', deshabilitar completamente

        Para reportes colombianos se habilita automáticamente.
        """
        if previous_options is None:
            previous_options = {}

        # Habilitar para reportes colombianos o si el filtro está configurado
        is_colombian = self._is_colombian_report()
        should_enable = is_colombian or self.filter_hide_partners_no_movement != 'never'

        if should_enable:
            previous_val = previous_options.get('hide_partners_no_movement')
            if previous_val is not None:
                options['hide_partners_no_movement'] = previous_val
            else:
                options['hide_partners_no_movement'] = self.filter_hide_partners_no_movement == 'by_default'
        else:
            options['hide_partners_no_movement'] = False

    def _init_options_hide_accounts_no_movement(self, options, previous_options=None):
        """
        Inicializa la opción para ocultar cuentas sin movimiento.

        Sigue el patrón exacto de _init_options_hide_0_lines del código nativo:
        1. Verificar si el filtro está habilitado (filter_* != 'never')
        2. Usar valor previo si existe, sino usar default del campo
        3. Si filter_* == 'never', deshabilitar completamente

        Este filtro es útil para el Libro Auxiliar, ocultando cuentas que solo
        tienen saldo inicial pero no movimientos en el período.

        Para reportes colombianos se habilita automáticamente.
        """
        if previous_options is None:
            previous_options = {}

        # Habilitar para reportes colombianos o si el filtro está configurado
        is_colombian = self._is_colombian_report()
        should_enable = is_colombian or self.filter_hide_accounts_no_movement != 'never'

        if should_enable:
            previous_val = previous_options.get('hide_accounts_no_movement')
            if previous_val is not None:
                options['hide_accounts_no_movement'] = previous_val
            else:
                options['hide_accounts_no_movement'] = self.filter_hide_accounts_no_movement == 'by_default'
        else:
            options['hide_accounts_no_movement'] = False

    def _init_options_color_formatting(self, options, previous_options=None):
        """
        Inicializa las opciones de formato de colores para el reporte.

        Modos de color disponibles:
        - accounting: Rojo para negativos, Verde para positivos
        - financial: Rojo para débitos, Azul para créditos
        - simple: Solo rojo para negativos

        Para reportes colombianos se habilita automáticamente.
        """
        if previous_options is None:
            previous_options = {}

        # Habilitar para reportes colombianos o si el filtro está configurado
        is_colombian = self._is_colombian_report()
        should_enable = is_colombian or self.filter_color_formatting != 'never'

        if should_enable:
            previous_val = previous_options.get('use_color_formatting')
            if previous_val is not None:
                options['use_color_formatting'] = previous_val
            else:
                options['use_color_formatting'] = self.filter_color_formatting == 'by_default'

            # Modo de color
            options['color_mode'] = previous_options.get('color_mode', 'accounting')

            # Opciones de modo de color
            options['color_modes'] = [
                {'id': 'accounting', 'name': _('Contable (Rojo=Negativo, Verde=Positivo)'), 'selected': False},
                {'id': 'financial', 'name': _('Financiero (Rojo=Débito, Azul=Crédito)'), 'selected': False},
                {'id': 'simple', 'name': _('Simple (Solo Rojo para Negativos)'), 'selected': False},
            ]
            for mode in options['color_modes']:
                mode['selected'] = (mode['id'] == options['color_mode'])

            # Usar símbolos contables (▲▼)
            options['use_accounting_symbols'] = previous_options.get('use_accounting_symbols', False)
        else:
            options['use_color_formatting'] = False
            options['color_mode'] = 'accounting'
            options['color_modes'] = []
            options['use_accounting_symbols'] = False

    def _init_options_secondary_currency(self, options, previous_options=None):
        """
        Inicializa las opciones de moneda secundaria.

        Cuando está habilitado, agrega columnas adicionales con valores
        convertidos a la moneda secundaria seleccionada.

        Para reportes colombianos se habilita automáticamente.
        """
        if previous_options is None:
            previous_options = {}

        # Habilitar para reportes colombianos o si el filtro está configurado
        is_colombian = self._is_colombian_report()
        should_enable = is_colombian or self.filter_secondary_currency

        if should_enable:
            options['show_secondary_currency'] = previous_options.get('show_secondary_currency', False)
            options['secondary_currency_id'] = previous_options.get('secondary_currency_id', False)

            # Obtener monedas activas para el selector
            company_currency = self.env.company.currency_id
            active_currencies = self.env['res.currency'].search([
                ('active', '=', True),
                ('id', '!=', company_currency.id),
            ])
            options['available_currencies'] = [
                {'id': c.id, 'name': c.name, 'symbol': c.symbol}
                for c in active_currencies
            ]
        else:
            options['show_secondary_currency'] = False
            options['secondary_currency_id'] = False
            options['available_currencies'] = []

    def _init_options_initial_balance_date(self, options, previous_options=None):
        """
        Inicializa la opción de mostrar fecha del último movimiento en saldos iniciales.
        """
        if previous_options is None:
            previous_options = {}

        if self.filter_initial_balance_date:
            options['show_initial_balance_date'] = previous_options.get('show_initial_balance_date', False)
        else:
            options['show_initial_balance_date'] = False

    def _init_options_export_wizard_button(self, options, previous_options=None):
        """
        Agrega el botón para abrir el wizard de exportación avanzado.

        NOTA: Este botón se desactivó temporalmente porque requiere un
        patrón especial de acción en el frontend. El wizard se puede
        acceder desde Exportar Avanzado en el menú de acciones.
        """
        # Por ahora desactivado - el wizard se accede desde el menú
        pass

    def action_open_export_wizard(self, options):
        """
        Abre el wizard de exportación avanzada con las opciones actuales.
        """
        import json
        wizard = self.env['libros.contables.export.wizard'].create({
            'report_id': self.id,
            'report_options': json.dumps(options),
            'doc_name': self.name,
        })
        return {
            'type': 'ir.actions.act_window',
            'name': _('Exportar Reporte'),
            'res_model': 'libros.contables.export.wizard',
            'view_mode': 'form',
            'res_id': wizard.id,
            'target': 'new',
            'context': {'dialog_size': 'medium'},
        }

    def _init_options_puc_hierarchy(self, options, previous_options=None):
        """
        Inicializa la opción de jerarquía PUC colombiano.

        Estructura de jerarquía:
        - Clase: 1 dígito (ej: 1 Activo, 2 Pasivo, 3 Patrimonio...)
        - Grupo: 2 dígitos (ej: 11 Disponible, 13 Deudores)
        - Cuenta: 4 dígitos (ej: 1105 Caja, 1110 Bancos)
        - Subcuenta: 6 dígitos (ej: 110505 Caja General)
        - Auxiliar: 8+ dígitos (ej: 11050501 Caja General ME)
        - Tercero: Agrupación por socio/tercero
        - Movimiento: Detalle de asientos contables

        Para reportes colombianos se habilita automáticamente.
        """
        if previous_options is None:
            previous_options = {}

        # Habilitar para reportes colombianos o si el filtro está configurado
        is_colombian = self._is_colombian_report()
        should_enable = is_colombian or self.filter_puc_hierarchy != 'never'

        if should_enable:
            previous_val = previous_options.get('puc_hierarchy')
            if previous_val is not None:
                options['puc_hierarchy'] = previous_val
            else:
                options['puc_hierarchy'] = self.filter_puc_hierarchy == 'by_default'

            # Opciones de nivel de jerarquía (hasta dónde mostrar)
            options['puc_hierarchy_levels'] = [
                {'id': 'all', 'name': _('Todos los niveles'), 'selected': False},
                {'id': 'clase', 'name': _('Clase (1 dígito)'), 'selected': False},
                {'id': 'grupo', 'name': _('Grupo (2 dígitos)'), 'selected': False},
                {'id': 'cuenta', 'name': _('Cuenta (4 dígitos)'), 'selected': False},
                {'id': 'subcuenta', 'name': _('Subcuenta (6 dígitos)'), 'selected': False},
                {'id': 'auxiliar', 'name': _('Auxiliar (8+ dígitos)'), 'selected': False},
                {'id': 'tercero', 'name': _('Tercero'), 'selected': False},
                {'id': 'movimiento', 'name': _('Movimientos'), 'selected': False},
            ]

            # Recuperar nivel seleccionado
            selected_level = previous_options.get('puc_hierarchy_level', 'all')
            options['puc_hierarchy_level'] = selected_level

            for level in options['puc_hierarchy_levels']:
                level['selected'] = (level['id'] == selected_level)

            # Opciones de modo de visualización
            options['puc_display_modes'] = [
                {'id': 'expandable', 'name': _('Expandible'), 'selected': True},
                {'id': 'flat', 'name': _('Vista Plana'), 'selected': False},
            ]
            selected_mode = previous_options.get('puc_display_mode', 'expandable')
            options['puc_display_mode'] = selected_mode
            for mode in options['puc_display_modes']:
                mode['selected'] = (mode['id'] == selected_mode)

            # Opción para mostrar terceros
            options['puc_show_partners'] = previous_options.get('puc_show_partners', True)

            # Opción para mostrar movimientos
            options['puc_show_movements'] = previous_options.get('puc_show_movements', True)
        else:
            options['puc_hierarchy'] = False
            options['puc_hierarchy_level'] = 'all'
            options['puc_hierarchy_levels'] = []
            options['puc_display_mode'] = 'expandable'
            options['puc_display_modes'] = []
            options['puc_show_partners'] = True
            options['puc_show_movements'] = True

    def _init_options_account_range(self, options, previous_options=None):
        """
        Inicializa el filtro por rango de cuentas contables.

        Sigue el patrón de _init_options_search_bar del código nativo:
        1. Verificar si el filtro está habilitado
        2. Marcar opción como habilitada
        3. Recuperar valores previos

        Para reportes colombianos se habilita automáticamente.
        """
        if previous_options is None:
            previous_options = {}

        # Habilitar para reportes colombianos o si el filtro está configurado
        is_colombian = self._is_colombian_report()
        should_enable = is_colombian or self.filter_account_range

        if not should_enable:
            return

        # Marcar que el filtro está habilitado (análogo a options['search_bar'] = True)
        options['account_range'] = True

        # Recuperar valores previos del filtro
        options['account_from'] = previous_options.get('account_from', '')
        options['account_to'] = previous_options.get('account_to', '')
        options['account_exclude'] = previous_options.get('account_exclude', [])

        # Calcular nombres para mostrar en UI
        self._compute_account_range_display_names(options)

    def _compute_account_range_display_names(self, options):
        """Calcula los nombres de cuentas para mostrar en la UI."""
        Account = self.env['account.account']

        if options.get('account_from'):
            account_from = Account.search([
                ('code', '=like', options['account_from'] + '%')
            ], limit=1, order='code')
            options['selected_account_from_name'] = (
                f"{account_from.code} - {account_from.name}"
                if account_from else options['account_from']
            )
        else:
            options['selected_account_from_name'] = ''

        if options.get('account_to'):
            account_to = Account.search([
                ('code', '=like', options['account_to'] + '%')
            ], limit=1, order='code')
            options['selected_account_to_name'] = (
                f"{account_to.code} - {account_to.name}"
                if account_to else options['account_to']
            )
        else:
            options['selected_account_to_name'] = ''

        if options.get('account_exclude'):
            excluded_accounts = Account.search([
                ('code', 'in', options['account_exclude'])
            ])
            options['selected_account_exclude_names'] = [
                f"{a.code} - {a.name}" for a in excluded_accounts
            ]
        else:
            options['selected_account_exclude_names'] = []

    def _init_options_account(self, options, previous_options=None):
        """
        Inicializa el filtro por cuentas contables.

        Sigue el patrón exacto de _init_options_partner del código nativo:
        1. Verificar si el filtro está habilitado
        2. Marcar options['account'] = True
        3. Recuperar IDs de cuentas previas
        4. Validar que las cuentas existen y el usuario tiene acceso
        5. Guardar nombres para mostrar en UI
        """
        if previous_options is None:
            previous_options = {}

        if not self.filter_account:
            return

        options['account'] = True

        # Recuperar IDs de cuentas previas
        previous_account_ids = previous_options.get('account_ids') or []
        selected_account_ids = [int(account) for account in previous_account_ids]

        # Buscar cuentas (search en lugar de browse para aplicar record rules)
        selected_accounts = (
            selected_account_ids and
            self.env['account.account'].with_context(active_test=False).search([
                ('id', 'in', selected_account_ids)
            ]) or self.env['account.account']
        )

        # Guardar nombres para mostrar en UI y IDs para el dominio
        options['selected_account_ids'] = selected_accounts.mapped('display_name')
        options['account_ids'] = selected_accounts.ids

    def _init_options_tax(self, options, previous_options=None):
        """
        Inicializa el filtro por impuestos.

        Sigue el patrón exacto de _init_options_partner del código nativo:
        1. Verificar si el filtro está habilitado
        2. Marcar options['tax'] = True
        3. Recuperar IDs de impuestos previos
        4. Validar que los impuestos existen y el usuario tiene acceso
        5. Guardar nombres para mostrar en UI
        6. Filtrar impuestos según el tipo de reporte (IVA, ReteFuente, etc.)
        """
        if previous_options is None:
            previous_options = {}

        if not self.filter_tax:
            return

        options['tax'] = True

        # Obtener dominio base de impuestos para el filtro
        tax_domain = self._get_tax_filter_domain()
        options['tax_filter_domain'] = tax_domain

        # Recuperar IDs de impuestos previos
        previous_tax_ids = previous_options.get('tax_ids') or []
        selected_tax_ids = [int(tax) for tax in previous_tax_ids]

        # Buscar impuestos aplicando el dominio del tipo de reporte
        if selected_tax_ids:
            # Validar que los impuestos previos siguen siendo válidos con el nuevo dominio
            valid_domain = [('id', 'in', selected_tax_ids)]
            if tax_domain:
                valid_domain += tax_domain
            selected_taxes = self.env['account.tax'].with_context(active_test=False).search(valid_domain)
        else:
            selected_taxes = self.env['account.tax']

        # Guardar nombres para mostrar en UI y IDs para el dominio
        options['selected_tax_ids'] = selected_taxes.mapped('display_name')
        options['tax_ids'] = selected_taxes.ids

    def _get_tax_filter_domain(self):
        """
        Obtiene el dominio para filtrar impuestos según el tipo de reporte.

        Para reportes específicos (IVA, ReteFuente, etc.), filtra solo
        los impuestos que corresponden a ese tipo.
        """
        # Verificar si el reporte tiene un handler personalizado con tipos de concepto
        if not self.custom_handler_model_id:
            return []

        handler_model = self.custom_handler_model_id.model
        if handler_model not in self.env:
            return []

        handler = self.env[handler_model]
        if not hasattr(handler, '_get_concept_types'):
            return []

        concept_types = handler._get_concept_types()
        if not concept_types:
            return []

        # Construir dominio para filtrar por tipo de impuesto colombiano
        # Prioriza l10n_co_tax_type del impuesto, sino busca por concepto de retención
        domain = ['|',
            ('l10n_co_tax_type', 'in', concept_types),
            ('withholding_concept_id.concept_type', 'in', concept_types),
        ]
        return domain

    def _init_options_withholding_filters(self, options, previous_options=None):
        """
        Inicializa los filtros de retención/impuestos colombianos.
        """
        if previous_options is None:
            previous_options = {}

        if not self.filter_withholding_filters:
            return

        options['show_withholding_filters'] = True

        # Inicializar tipos de operación
        if self.filter_operation_type:
            options['operation_types'] = [
                {'id': 'all', 'name': _('Todas'), 'selected': True},
                {'id': 'purchase', 'name': _('Compras'), 'selected': False},
                {'id': 'sale', 'name': _('Ventas'), 'selected': False},
                {'id': 'purchase_refund', 'name': _('NC Proveedor'), 'selected': False},
                {'id': 'sale_refund', 'name': _('NC Cliente'), 'selected': False},
            ]
            options['operation_type_filter'] = previous_options.get('operation_type_filter', 'all')
            # Actualizar selección
            for ot in options['operation_types']:
                ot['selected'] = ot['id'] == options['operation_type_filter']

        # Inicializar tipos de concepto
        if self.filter_concept_type:
            options['concept_types'] = [
                {'id': 'all', 'name': _('Todos'), 'selected': True},
                {'id': 'retefuente', 'name': _('ReteFuente'), 'selected': False},
                {'id': 'reteiva', 'name': _('ReteIVA'), 'selected': False},
                {'id': 'reteica', 'name': _('ReteICA'), 'selected': False},
                {'id': 'iva', 'name': _('IVA'), 'selected': False},
                {'id': 'inc', 'name': _('INC'), 'selected': False},
            ]
            options['concept_type_filter'] = previous_options.get('concept_type_filter', 'all')
            for ct in options['concept_types']:
                ct['selected'] = ct['id'] == options['concept_type_filter']

        # Inicializar conceptos específicos de retención
        if self.filter_withholding_concept:
            options['withholding_concepts'] = self._get_withholding_concepts()
            options['withholding_concept_ids'] = previous_options.get('withholding_concept_ids', [])

        # Inicializar filtro de ciudades
        if self.filter_city:
            options['cities'] = self._get_cities_for_filter()
            options['city_ids'] = previous_options.get('city_ids', [])
            # Actualizar selección
            for city in options.get('cities', []):
                city['selected'] = city['id'] in options['city_ids']

    def _get_cities_for_filter(self):
        """Obtiene las ciudades disponibles desde los partners con movimientos de impuestos."""
        query = """
            SELECT DISTINCT rc.id, rc.name, rcs.name as state_name
            FROM res_city rc
            JOIN res_partner rp ON rp.city_id = rc.id
            JOIN account_move_line aml ON aml.partner_id = rp.id
            LEFT JOIN res_country_state rcs ON rcs.id = rc.state_id
            WHERE aml.tax_line_id IS NOT NULL
            ORDER BY rcs.name, rc.name
        """
        self._cr.execute(query)

        cities = []
        for row in self._cr.dictfetchall():
            state_name = row.get('state_name', '')
            city_name = row.get('name', '')
            display_name = f"{city_name} ({state_name})" if state_name else city_name

            cities.append({
                'id': row['id'],
                'name': display_name,
                'selected': False,
            })

        return cities

    def _get_withholding_concepts(self):
        """Obtiene los conceptos de retención desde tax_base_threshold."""
        # Buscar impuestos con configuración de base mínima
        taxes = self.env['account.tax'].search([
            ('type_tax_use', 'in', ['purchase', 'sale']),
            ('amount_type', '=', 'percent'),
        ], order='name')

        concepts = []
        for tax in taxes:
            concepts.append({
                'id': tax.id,
                'code': tax.name.split()[0] if tax.name else str(tax.id),
                'name': tax.name,
                'selected': False,
            })
        return concepts

    # =========================================================================
    # MÉTODOS _get_options_*_domain (para generar dominios de búsqueda)
    # =========================================================================

    @api.model
    def _get_options_account_range_domain(self, options):
        """
        Genera el dominio para filtrar por rango de cuentas.

        Sigue el patrón de _get_options_partner_domain del código nativo.
        """
        domain = []

        if not options.get('account_range'):
            return domain

        account_from = options.get('account_from', '')
        account_to = options.get('account_to', '')

        if account_from:
            domain.append(('account_id.code', '>=', account_from))

        if account_to:
            # Agregar sufijo 'z' para incluir todas las subcuentas
            domain.append(('account_id.code', '<=', account_to + 'z'))

        account_exclude = options.get('account_exclude', [])
        if account_exclude:
            domain.append(('account_id.code', 'not in', account_exclude))

        return domain

    @api.model
    def _get_options_account_domain(self, options):
        """
        Genera el dominio para filtrar por cuentas contables específicas.

        Sigue el patrón exacto de _get_options_partner_domain del código nativo.
        """
        domain = []
        if options.get('account_ids'):
            account_ids = [int(account) for account in options['account_ids']]
            domain.append(('account_id', 'in', account_ids))
        return domain

    @api.model
    def _get_options_tax_domain(self, options):
        """
        Genera el dominio para filtrar por impuestos específicos.

        Sigue el patrón exacto de _get_options_partner_domain del código nativo.
        Filtra las líneas de asiento que tienen los impuestos seleccionados.
        """
        domain = []
        if options.get('tax_ids'):
            tax_ids = [int(tax) for tax in options['tax_ids']]
            domain.append(('tax_ids', 'in', tax_ids))
        return domain

    # =========================================================================
    # Override _get_options_domain para incluir filtros personalizados
    # =========================================================================

    def _get_options_domain(self, options, date_scope):
        """
        Extiende el método nativo para incluir dominios de filtros personalizados.

        Este método es llamado por _get_report_query para construir el dominio
        completo de búsqueda. Agregamos aquí nuestros filtros personalizados
        para que sean aplicados automáticamente en todas las queries.
        """
        # Llamar al método original
        domain = super()._get_options_domain(options, date_scope)

        # Agregar dominio de rango de cuentas
        domain += self._get_options_account_range_domain(options)

        # Agregar dominio de cuentas específicas (si se usa filter_account)
        domain += self._get_options_account_domain(options)

        # Agregar dominio de impuestos (si se usa filter_tax)
        domain += self._get_options_tax_domain(options)

        return domain

    # =========================================================================
    # FIX: Bug en Odoo 18 Enterprise - IndexError en comparison periods
    # =========================================================================

    def _adjust_date_for_joined_comparison(self, options, period_date_from):
        """
        Override para corregir bug en Odoo 18 Enterprise.

        El código nativo falla cuando options['comparison']['periods'] es lista vacía:
        options['comparison'].get('periods', [{}])[-1].get('date_from')

        IndexError: list index out of range

        Este fix verifica que periods tenga elementos antes de acceder al índice.
        """
        import datetime
        comparison_filter = options.get('comparison', {}).get('filter')
        if comparison_filter == 'previous_period':
            periods = options.get('comparison', {}).get('periods', [])
            # FIX: Verificar que periods no esté vacía
            if periods and periods[-1].get('date_from'):
                comparison_date_from = datetime.datetime.strptime(
                    periods[-1].get('date_from'), '%Y-%m-%d'
                )
                return min(period_date_from, comparison_date_from)
        return period_date_from

    # =========================================================================
    # get_report_information (Paso 3 del patrón nativo)
    # Exponer filtros al frontend en el dict 'filters'
    # =========================================================================

    def get_report_information(self, options):
        """
        Extiende la información del reporte para incluir filtros personalizados.

        Sigue el patrón nativo de account_reports:
        - Los filtros se exponen en result['filters'] con claves 'show_*'
        - El frontend lee estos valores en this.controller.filters

        Para reportes colombianos (que usan handlers del módulo libros_contables_colombia),
        los filtros se habilitan automáticamente sin necesidad de configuración manual.
        """
        _logger.info("=== get_report_information ===")
        _logger.info("  Report: %s (ID: %s)", self.name, self.id)

        result = super().get_report_information(options)

        # Detectar si es un reporte colombiano para habilitar filtros automáticamente
        is_colombian = self._is_colombian_report()
        _logger.info("  Is Colombian report: %s", is_colombian)

        # =====================================================================
        # FILTROS GLOBALES PARA REPORTES COLOMBIANOS
        # Si es un reporte colombiano, habilitar filtros automáticamente
        # =====================================================================

        if is_colombian:
            # Formato de colores - siempre disponible en reportes colombianos
            result['filters']['show_color_formatting'] = 'optional'

            # Moneda secundaria - siempre disponible
            result['filters']['show_secondary_currency'] = True

            # Jerarquía PUC - siempre disponible en reportes colombianos
            result['filters']['show_puc_hierarchy'] = 'optional'

            # Rango de cuentas - siempre disponible
            result['filters']['show_account_range'] = True

            # Ocultar cuentas/terceros sin movimiento - disponible por defecto
            result['filters']['show_hide_accounts_no_movement'] = 'optional'
            result['filters']['show_hide_partners_no_movement'] = 'optional'

        # =====================================================================
        # FILTROS ESPECÍFICOS POR CONFIGURACIÓN DEL REPORTE
        # Estos se habilitan solo si el campo está configurado en el reporte
        # =====================================================================

        # Filtro de terceros sin movimiento (configuración explícita)
        if self.filter_hide_partners_no_movement != 'never':
            result['filters']['show_hide_partners_no_movement'] = self.filter_hide_partners_no_movement

        # Filtro de cuentas sin movimiento (configuración explícita)
        if self.filter_hide_accounts_no_movement != 'never':
            result['filters']['show_hide_accounts_no_movement'] = self.filter_hide_accounts_no_movement

        # Filtro de jerarquía PUC (configuración explícita)
        if self.filter_puc_hierarchy != 'never':
            result['filters']['show_puc_hierarchy'] = self.filter_puc_hierarchy

        # Filtro de rango de cuentas
        if self.filter_account_range:
            result['filters']['show_account_range'] = True

        # Filtro de cuentas específicas (selector múltiple)
        if self.filter_account:
            result['filters']['show_account'] = True

        # Filtro de impuestos (selector múltiple)
        if self.filter_tax:
            result['filters']['show_tax'] = True

        # Filtros de retención colombianos
        if self.filter_withholding_filters:
            result['filters']['show_withholding_filters'] = True
        if self.filter_operation_type:
            result['filters']['show_operation_type'] = True
        if self.filter_concept_type:
            result['filters']['show_concept_type'] = True
        if self.filter_withholding_concept:
            result['filters']['show_withholding_concept'] = True
        if self.filter_person_type:
            result['filters']['show_person_type'] = True
        if self.filter_product_type:
            result['filters']['show_product_type'] = True
        if self.filter_city:
            result['filters']['show_city'] = True

        # Filtro de formato de colores (configuración explícita)
        if self.filter_color_formatting != 'never':
            result['filters']['show_color_formatting'] = self.filter_color_formatting

        # Filtro de moneda secundaria
        if self.filter_secondary_currency:
            result['filters']['show_secondary_currency'] = True

        # Filtro de fecha de último movimiento en saldo inicial
        if self.filter_initial_balance_date:
            result['filters']['show_initial_balance_date'] = True

        _logger.info("  Result filters: %s", list(result.get('filters', {}).keys()))

        return result

    # =========================================================================
    # MÉTODOS AUXILIARES
    # =========================================================================

    def _is_trial_balance_partner_report(self, options=None):
        """Verifica si es el Balance de Prueba por Tercero."""
        if self.custom_handler_model_id:
            return self.custom_handler_model_id.model == 'account.trial.balance.partner.report.handler'
        return False

    def _is_trial_balance_puc_report(self, options=None):
        """Verifica si es el Balance de Prueba PUC."""
        if self.custom_handler_model_id:
            return self.custom_handler_model_id.model == 'account.trial.balance.puc.report.handler'
        return False

    def _is_auxiliary_book_report(self, options):
        """Verifica si es el Libro Auxiliar (normal o analítico)."""
        if self.custom_handler_model_id:
            return self.custom_handler_model_id.model == 'account.auxiliary.book.colombia.handler'
        return False

    def _is_auxiliary_book_analytic(self, options):
        """Verifica si es el Libro Auxiliar Analítico."""
        name_lower = (self.name or '').lower()
        return self._is_auxiliary_book_report(options) and ('analytic' in name_lower or 'analítico' in name_lower)

    def _is_colombian_report(self):
        """
        Verifica si este reporte usa un handler colombiano del módulo libros_contables_colombia.

        Returns:
            bool: True si el reporte usa un handler colombiano
        """
        if not self.custom_handler_model_id:
            return False
        return self.custom_handler_model_id.model in COLOMBIAN_HANDLER_MODELS

    # =========================================================================
    # FORMATO DE COLORES Y SÍMBOLOS CONTABLES
    # Override de _build_column_dict para aplicar colores y símbolos
    # =========================================================================

    def _build_column_dict(self, col_value, col_data, options=None, currency=False, digits=1,
                           column_expression=None, has_sublines=False, report_line_id=None):
        """
        Override para agregar formato de colores y símbolos contables.

        Cuando use_color_formatting está habilitado:
        - accounting: Verde (▲) para positivos, Rojo (▼) para negativos
        - financial: Rojo para débitos, Azul para créditos
        - simple: Solo rojo para negativos

        Cuando use_accounting_symbols está habilitado:
        - Agrega ▲ para valores positivos
        - Agrega ▼ para valores negativos
        """
        # Llamar al método original
        result = super()._build_column_dict(
            col_value, col_data, options=options, currency=currency, digits=digits,
            column_expression=column_expression, has_sublines=has_sublines, report_line_id=report_line_id
        )

        if not result:
            return result

        options = options or {}

        # Solo aplicar formato si está habilitado y es un valor numérico
        use_color = options.get('use_color_formatting', False)
        use_symbols = options.get('use_accounting_symbols', False)
        color_mode = options.get('color_mode', 'accounting')

        if not (use_color or use_symbols):
            return result

        # Solo procesar valores numéricos
        figure_type = result.get('figure_type', 'string')
        if figure_type not in ('monetary', 'float', 'integer', 'percentage'):
            return result

        no_format = result.get('no_format')
        if no_format is None:
            return result

        try:
            value = float(no_format)
        except (ValueError, TypeError):
            return result

        # Obtener la etiqueta de expresión para determinar el tipo de columna
        expr_label = result.get('expression_label', '')

        # Determinar clase CSS y símbolo según el modo de color
        css_class = result.get('class', '')
        symbol = ''

        if use_color:
            if color_mode == 'accounting':
                # Contable: Verde=Positivo, Rojo=Negativo
                if value > 0:
                    css_class = (css_class + ' text-success').strip()
                elif value < 0:
                    css_class = (css_class + ' text-danger').strip()

            elif color_mode == 'financial':
                # Financiero: Rojo=Débito, Azul=Crédito
                if expr_label in ('debit', 'Débito', 'debito'):
                    if value != 0:
                        css_class = (css_class + ' text-danger').strip()
                elif expr_label in ('credit', 'Crédito', 'credito'):
                    if value != 0:
                        css_class = (css_class + ' text-primary').strip()
                elif value < 0:
                    # Para otras columnas, rojo si negativo
                    css_class = (css_class + ' text-danger').strip()

            elif color_mode == 'simple':
                # Simple: Solo rojo para negativos
                if value < 0:
                    css_class = (css_class + ' text-danger').strip()

        if use_symbols:
            if value > 0:
                symbol = '▲ '
            elif value < 0:
                symbol = '▼ '

        # Aplicar clase CSS
        if css_class:
            result['class'] = css_class

        # Aplicar símbolo al valor formateado
        if symbol:
            if 'name' in result and result['name']:
                # Si ya hay un nombre formateado, agregar símbolo
                result['name'] = f"{symbol}{result['name']}"
            else:
                # Formatear el valor manualmente con el símbolo
                formatted_value = self._format_value_with_symbol(value, figure_type, symbol, options, currency)
                if formatted_value:
                    result['name'] = formatted_value

        return result

    def _format_value_with_symbol(self, value, figure_type, symbol, options, currency=False):
        """
        Formatea un valor numérico con símbolo contable (▲▼).

        Args:
            value: Valor numérico
            figure_type: Tipo de figura ('monetary', 'float', etc.)
            symbol: Símbolo a agregar ('▲ ' o '▼ ')
            options: Opciones del reporte
            currency: Moneda para formato

        Returns:
            str: Valor formateado con símbolo o None
        """
        from odoo.tools.misc import formatLang

        try:
            if figure_type == 'monetary':
                # Obtener la moneda
                if currency:
                    currency_obj = currency
                else:
                    currency_obj = self.env.company.currency_id

                # Formatear como moneda
                formatted = formatLang(
                    self.env,
                    abs(value),  # Usar valor absoluto, el símbolo indica el signo
                    currency_obj=currency_obj
                )

                # Agregar símbolo y signo
                if value < 0:
                    return f"{symbol}-{formatted}"
                else:
                    return f"{symbol}{formatted}"

            elif figure_type == 'percentage':
                formatted = f"{abs(value):.2f}%"
                if value < 0:
                    return f"{symbol}-{formatted}"
                else:
                    return f"{symbol}{formatted}"

            elif figure_type in ('float', 'integer'):
                if figure_type == 'integer':
                    formatted = f"{int(abs(value)):,}"
                else:
                    formatted = f"{abs(value):,.2f}"

                if value < 0:
                    return f"{symbol}-{formatted}"
                else:
                    return f"{symbol}{formatted}"

        except Exception:
            pass

        return None

    def _get_trial_balance_partner_pdf_html(self, options, lines):
        """
        Genera HTML autocontenido para el Balance de Prueba por Tercero.
        Sin dependencias externas para evitar errores de wkhtmltopdf.
        """
        company = self.env.company

        # Logo como base64 data URI
        logo_html = ''
        if company.logo:
            logo_b64 = base64.b64encode(company.logo).decode('utf-8')
            logo_html = f'<img src="data:image/png;base64,{logo_b64}" style="max-height: 45px; max-width: 90px;" alt="Logo"/>'

        # Fechas
        date_from = options.get('date', {}).get('date_from', '')
        date_to = options.get('date', {}).get('date_to', '')

        # Nombre de empresa desde options
        company_names = []
        for c in options.get('companies', []):
            company_names.append(c.get('name', ''))
        company_name = ', '.join(company_names) if company_names else company.name

        def format_value(col):
            """Formatea el valor de una columna para mostrar en el PDF."""
            if col is None:
                return ''
            val = col.get('no_format')
            if val is None:
                return ''
            figure_type = col.get('figure_type', 'string')
            if figure_type == 'monetary':
                try:
                    return f"$ {float(val):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                except (ValueError, TypeError):
                    return str(val) if val else ''
            elif figure_type == 'date':
                if val:
                    from datetime import date as dt_date
                    if isinstance(val, dt_date):
                        return val.strftime('%d/%m/%Y')
                    return str(val)
                return ''
            return str(val) if val else ''

        def get_col_by_label(cols, label):
            """Obtiene una columna por su expression_label."""
            for col in cols:
                if col.get('expression_label') == label:
                    return col
            return None

        # Generar filas de la tabla
        # Orden de columnas: CÓDIGO, CUENTA, TERCERO, NIT, ÚLT.MOV, SDO.INICIAL, DÉBITOS, CRÉDITOS, SDO.FINAL
        rows_html = []
        for line in lines:
            lvl = line.get('level', 1)
            is_total = 'total' in str(line.get('id', ''))
            cols = line.get('columns', [])

            # Obtener valores de columnas en el orden correcto
            # Para nivel 5 (terceros): la columna "Nombre Cuenta" debe estar vacía
            # ya que el nombre del tercero va en su propia columna
            account_name_val = format_value(get_col_by_label(cols, 'account_name'))
            partner_name_val = format_value(get_col_by_label(cols, 'partner_name'))

            # Si es nivel 5 (tercero), limpiar nombre de cuenta para evitar duplicación
            if lvl == 5 or lvl == 6:
                account_name_val = ''

            col_vals = [
                format_value(get_col_by_label(cols, 'account_code')),
                account_name_val,
                partner_name_val,
                format_value(get_col_by_label(cols, 'partner_vat')),
                format_value(get_col_by_label(cols, 'last_move_date')),
                format_value(get_col_by_label(cols, 'initial_balance')),
                format_value(get_col_by_label(cols, 'debit')),
                format_value(get_col_by_label(cols, 'credit')),
                format_value(get_col_by_label(cols, 'final_balance')),
            ]

            # Estilos según nivel - tonos más claros y sutiles con texto negro legible
            if is_total:
                row_style = 'background-color: #2c3e50; color: #fff; font-weight: bold; font-size: 8pt;'
                cell_border = 'border: 1px solid #2c3e50;'
                padding_left = '3px'
            elif lvl == 1:
                # Nivel 1 (Clase): Azul oscuro suave
                row_style = 'background-color: #34495e; color: #fff; font-weight: bold;'
                cell_border = 'border-bottom: 1px solid #2c3e50;'
                padding_left = '2px'
            elif lvl == 2:
                # Nivel 2 (Grupo): Gris azulado medio
                row_style = 'background-color: #7f8c8d; color: #fff; font-weight: bold;'
                cell_border = 'border-bottom: 1px solid #6c7a7b;'
                padding_left = '8px'
            elif lvl == 3:
                # Nivel 3 (Cuenta): Gris claro con texto negro
                row_style = 'background-color: #bdc3c7; color: #2c3e50; font-weight: bold;'
                cell_border = 'border-bottom: 1px solid #a0a6aa;'
                padding_left = '16px'
            elif lvl == 4:
                # Nivel 4 (Subcuenta): Gris muy claro con texto negro
                row_style = 'background-color: #ecf0f1; color: #2c3e50;'
                cell_border = 'border-bottom: 1px solid #d5dbdb;'
                padding_left = '24px'
            else:  # nivel 5 (terceros)
                # Nivel 5 (Tercero): Fondo blanco con texto negro
                row_style = 'background-color: #fafafa; color: #333; font-size: 6.5pt;'
                cell_border = 'border-bottom: 1px solid #e8e8e8;'
                padding_left = '32px'

            # Construir fila
            row_html = f'<tr style="{row_style}">'
            for i, val in enumerate(col_vals):
                cell_style = f'padding: 2px; {cell_border}'
                if i == 0:  # Primera columna (código) - con indentación
                    cell_style += f' padding-left: {padding_left};'
                elif i >= 5:  # Columnas numéricas
                    cell_style += ' text-align: right;'
                elif i == 4:  # Última mov - centrado
                    cell_style += ' text-align: center;'
                elif i == 2 and lvl == 5:  # Tercero en cursiva
                    cell_style += ' font-style: italic;'

                row_html += f'<td style="{cell_style}">{val}</td>'
            row_html += '</tr>'
            rows_html.append(row_html)

        rows_content = '\n'.join(rows_html)

        # HTML completo autocontenido
        html = f'''<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8"/>
    <style>
        @page {{
            size: letter landscape;
            margin: 8mm 5mm 12mm 5mm;
        }}
        * {{
            box-sizing: border-box;
        }}
        body {{
            font-family: Arial, Helvetica, sans-serif;
            font-size: 8pt;
            line-height: 1.1;
            margin: 0;
            padding: 5px;
            -webkit-print-color-adjust: exact !important;
            print-color-adjust: exact !important;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
        }}
        .header-table {{
            margin-bottom: 3px;
            border-bottom: 2px solid #2c3e50;
        }}
        .data-table {{
            font-size: 7pt;
            table-layout: fixed;
        }}
        .data-table th {{
            background-color: #2c3e50;
            color: #fff;
            padding: 4px 3px;
            text-align: left;
            border: 1px solid #1a252f;
            font-weight: bold;
        }}
    </style>
</head>
<body>
    <!-- ENCABEZADO -->
    <table class="header-table">
        <tr>
            <td style="width: 12%; vertical-align: middle; padding: 2px;">
                {logo_html}
            </td>
            <td style="width: 60%; text-align: center; vertical-align: middle; padding: 2px;">
                <div style="font-weight: bold; font-size: 11pt;">{company_name}</div>
                <div style="font-size: 8pt;">NIT: {company.vat or ''}</div>
                <div style="font-weight: bold; font-size: 10pt; margin-top: 2px; text-transform: uppercase;">
                    BALANCE DE PRUEBA POR TERCERO
                </div>
                <div style="font-size: 8pt;">Del {date_from} al {date_to}</div>
            </td>
            <td style="width: 28%; text-align: right; vertical-align: middle; font-size: 7pt; padding: 2px;">
                <div>{company.street or ''}</div>
                <div>{company.city or ''} - {company.country_id.name if company.country_id else ''}</div>
                <div>{company.phone or ''}</div>
            </td>
        </tr>
    </table>

    <!-- TABLA PRINCIPAL -->
    <table class="data-table">
        <thead>
            <tr>
                <th style="width: 7%;">CÓDIGO</th>
                <th style="width: 20%;">NOMBRE CUENTA</th>
                <th style="width: 16%;">TERCERO</th>
                <th style="width: 9%;">NIT/CC</th>
                <th style="width: 6%; text-align: center;">ÚLT.MOV</th>
                <th style="width: 10.5%; text-align: right;">SALDO INICIAL</th>
                <th style="width: 10.5%; text-align: right;">DÉBITOS</th>
                <th style="width: 10.5%; text-align: right;">CRÉDITOS</th>
                <th style="width: 10.5%; text-align: right;">SALDO FINAL</th>
            </tr>
        </thead>
        <tbody>
            {rows_content}
        </tbody>
    </table>
</body>
</html>'''
        return html

    def _get_trial_balance_partner_footer_html(self, options):
        """Genera el HTML del pie de página con número de página.

        wkhtmltopdf pasa los valores page y topage como parámetros en la URL
        del documento, por lo que necesitamos JavaScript para leerlos.
        """
        company = self.env.company
        company_names = []
        for c in options.get('companies', []):
            company_names.append(c.get('name', ''))
        company_name = ', '.join(company_names) if company_names else company.name

        return f'''<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8"/>
    <script>
        function subst() {{
            var vars = {{}};
            var query_strings_from_url = document.location.search.substring(1).split('&');
            for (var query_string in query_strings_from_url) {{
                if (query_strings_from_url.hasOwnProperty(query_string)) {{
                    var temp_var = query_strings_from_url[query_string].split('=', 2);
                    vars[temp_var[0]] = decodeURIComponent(temp_var[1]);
                }}
            }}
            var page_elem = document.getElementById('page');
            var topage_elem = document.getElementById('topage');
            if (page_elem) page_elem.textContent = vars.page || '';
            if (topage_elem) topage_elem.textContent = vars.topage || '';
        }}
    </script>
    <style>
        body {{
            font-family: Arial, Helvetica, sans-serif;
            font-size: 7pt;
            margin: 0;
            padding: 0 10px;
        }}
        .footer-table {{
            width: 100%;
            border-top: 1px solid #2c3e50;
            padding-top: 3px;
        }}
        .footer-table td {{
            padding: 2px 5px;
        }}
    </style>
</head>
<body onload="subst()">
    <table class="footer-table">
        <tr>
            <td style="width: 33%; text-align: left; color: #555;">
                {company_name} | NIT: {company.vat or ''}
            </td>
            <td style="width: 34%; text-align: center; color: #555;">
                Balance de Prueba por Tercero
            </td>
            <td style="width: 33%; text-align: right; color: #333; font-weight: bold;">
                Página <span id="page"></span> de <span id="topage"></span>
            </td>
        </tr>
    </table>
</body>
</html>'''

    def export_to_pdf(self, options):
        """
        Sobrescribe para usar formato US Letter Landscape en reportes especiales.
        """
        self.ensure_one()

        # Verificar si es el reporte de Balance por Tercero
        if self._is_trial_balance_partner_report(options):
            return self._export_trial_balance_partner_pdf(options)

        # Verificar si es el reporte de Balance PUC
        if self._is_trial_balance_puc_report(options):
            return self._export_trial_balance_puc_pdf(options)

        # Verificar si es el Libro Auxiliar
        if self._is_auxiliary_book_report(options):
            return self._export_auxiliary_book_pdf(options)

        # Para otros reportes, usar el método estándar
        return super().export_to_pdf(options)

    def export_to_xlsx(self, options, response=None):
        """
        Sobrescribe para usar exportación Excel con outlines en reportes especiales.
        """
        self.ensure_one()

        # Verificar si es el reporte de Balance por Tercero
        if self._is_trial_balance_partner_report(options):
            return self._export_trial_balance_partner_xlsx(options)

        # Verificar si es el reporte de Balance PUC
        if self._is_trial_balance_puc_report(options):
            return self._export_trial_balance_puc_xlsx(options)

        # Verificar si es el Libro Auxiliar
        if self._is_auxiliary_book_report(options):
            return self._export_auxiliary_book_xlsx(options)

        # Para otros reportes, usar el método estándar
        return super().export_to_xlsx(options, response)

    def _export_trial_balance_partner_xlsx(self, options):
        """
        Genera Excel para Balance de Prueba por Tercero con agrupamiento por niveles PUC.
        Usa outlines de Excel para permitir expandir/contraer niveles.
        """
        if not xlsxwriter:
            raise ValueError("xlsxwriter no está instalado")

        print_options = self.get_options(previous_options={**options, 'export_mode': 'print'})

        # Obtener líneas expandidas
        handler_model = self.custom_handler_model_id.model
        handler = self.env[handler_model]
        lines = handler._get_lines_for_pdf(self, print_options)

        # Crear workbook
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {
            'in_memory': True,
            'strings_to_formulas': False,
        })

        # Configurar outline:
        # Parámetros de outline_settings según documentación xlsxwriter:
        # - visible=True: Outlines visibles (muestra los símbolos +/-)
        # - symbols_below=False: símbolos ARRIBA del grupo (no debajo)
        # - symbols_right=False: símbolos a la izquierda
        # - auto_style=False: sin estilos automáticos
        sheet = workbook.add_worksheet('Balance por Tercero')
        sheet.outline_settings(True, False, False, False)

        # Definir formatos
        formats = self._get_xlsx_formats(workbook)

        # Escribir encabezado
        row = self._write_xlsx_header(sheet, formats, print_options)

        # Escribir encabezados de columnas
        row = self._write_xlsx_column_headers(sheet, formats, row)

        # Escribir datos con outlines
        row = self._write_xlsx_data_with_outlines(sheet, formats, lines, row, print_options)

        # Ajustar anchos de columna
        self._set_xlsx_column_widths(sheet)

        workbook.close()
        output.seek(0)

        # Post-procesar con openpyxl para asegurar que showOutlineSymbols=True
        # xlsxwriter no tiene forma directa de configurar este atributo
        if openpyxl_load_workbook:
            try:
                wb = openpyxl_load_workbook(output)
                ws = wb.active
                for view in ws.views.sheetView:
                    view.showOutlineSymbols = True

                # Guardar a un nuevo BytesIO
                final_output = io.BytesIO()
                wb.save(final_output)
                final_output.seek(0)
                generated_file = final_output.read()
                final_output.close()
            except Exception as e:
                _logger.warning("Error al post-procesar Excel con openpyxl: %s", str(e))
                output.seek(0)
                generated_file = output.read()
        else:
            generated_file = output.read()

        output.close()

        return {
            'file_name': self.get_default_report_filename(options, 'xlsx'),
            'file_content': generated_file,
            'file_type': 'xlsx',
        }

    def _get_xlsx_formats(self, workbook):
        """Define los formatos para el Excel."""
        formats = {}

        # Formato de título
        formats['title'] = workbook.add_format({
            'bold': True,
            'font_size': 14,
            'align': 'center',
            'valign': 'vcenter',
        })

        # Formato de subtítulo
        formats['subtitle'] = workbook.add_format({
            'font_size': 10,
            'align': 'center',
            'valign': 'vcenter',
        })

        # Formato de encabezado de columna
        formats['header'] = workbook.add_format({
            'bold': True,
            'font_size': 9,
            'bg_color': '#2c3e50',
            'font_color': 'white',
            'align': 'center',
            'valign': 'vcenter',
            'border': 1,
            'text_wrap': True,
        })

        # Formato numérico para encabezado
        formats['header_right'] = workbook.add_format({
            'bold': True,
            'font_size': 9,
            'bg_color': '#2c3e50',
            'font_color': 'white',
            'align': 'right',
            'valign': 'vcenter',
            'border': 1,
        })

        # Nivel 1 - Clase (más oscuro)
        formats['level_1'] = workbook.add_format({
            'bold': True,
            'font_size': 10,
            'bg_color': '#34495e',
            'font_color': 'white',
            'border': 1,
        })
        formats['level_1_num'] = workbook.add_format({
            'bold': True,
            'font_size': 10,
            'bg_color': '#34495e',
            'font_color': 'white',
            'border': 1,
            'num_format': '#,##0.00',
            'align': 'right',
        })

        # Nivel 2 - Grupo
        formats['level_2'] = workbook.add_format({
            'bold': True,
            'font_size': 9,
            'bg_color': '#7f8c8d',
            'font_color': 'white',
            'border': 1,
        })
        formats['level_2_num'] = workbook.add_format({
            'bold': True,
            'font_size': 9,
            'bg_color': '#7f8c8d',
            'font_color': 'white',
            'border': 1,
            'num_format': '#,##0.00',
            'align': 'right',
        })

        # Nivel 3 - Cuenta
        formats['level_3'] = workbook.add_format({
            'bold': True,
            'font_size': 9,
            'bg_color': '#bdc3c7',
            'font_color': '#2c3e50',
            'border': 1,
        })
        formats['level_3_num'] = workbook.add_format({
            'bold': True,
            'font_size': 9,
            'bg_color': '#bdc3c7',
            'font_color': '#2c3e50',
            'border': 1,
            'num_format': '#,##0.00',
            'align': 'right',
        })

        # Nivel 4 - Subcuenta
        formats['level_4'] = workbook.add_format({
            'font_size': 9,
            'bg_color': '#ecf0f1',
            'font_color': '#2c3e50',
            'border': 1,
        })
        formats['level_4_num'] = workbook.add_format({
            'font_size': 9,
            'bg_color': '#ecf0f1',
            'font_color': '#2c3e50',
            'border': 1,
            'num_format': '#,##0.00',
            'align': 'right',
        })

        # Nivel 5 - Tercero
        formats['level_5'] = workbook.add_format({
            'font_size': 8,
            'bg_color': '#fafafa',
            'font_color': '#333333',
            'border': 1,
            'italic': True,
        })
        formats['level_5_num'] = workbook.add_format({
            'font_size': 8,
            'bg_color': '#fafafa',
            'font_color': '#333333',
            'border': 1,
            'num_format': '#,##0.00',
            'align': 'right',
        })

        # Total general
        formats['total'] = workbook.add_format({
            'bold': True,
            'font_size': 10,
            'bg_color': '#2c3e50',
            'font_color': 'white',
            'border': 2,
        })
        formats['total_num'] = workbook.add_format({
            'bold': True,
            'font_size': 10,
            'bg_color': '#2c3e50',
            'font_color': 'white',
            'border': 2,
            'num_format': '#,##0.00',
            'align': 'right',
        })

        # Formato fecha
        formats['date'] = workbook.add_format({
            'font_size': 8,
            'align': 'center',
            'num_format': 'dd/mm/yyyy',
        })

        # =========================================================================
        # FORMATOS CON COLORES (para use_color_formatting)
        # =========================================================================
        # Colores Bootstrap
        positive_color = '#198754'  # Verde
        negative_color = '#dc3545'  # Rojo
        zero_color = '#999999'      # Gris

        # Niveles con colores (5-6 son los niveles de detalle)
        for level in [4, 5, 6]:
            bg_color = '#fafafa' if level == 5 else '#ecf0f1' if level == 4 else '#ffffff'

            # Positivo (verde)
            formats[f'level_{level}_positive'] = workbook.add_format({
                'font_size': 8 if level >= 5 else 9,
                'bg_color': bg_color,
                'font_color': positive_color,
                'border': 1,
                'num_format': '#,##0.00',
                'align': 'right',
            })

            # Negativo (rojo con paréntesis)
            formats[f'level_{level}_negative'] = workbook.add_format({
                'font_size': 8 if level >= 5 else 9,
                'bg_color': bg_color,
                'font_color': negative_color,
                'border': 1,
                'num_format': '(#,##0.00)',
                'align': 'right',
            })

            # Cero (gris)
            formats[f'level_{level}_zero'] = workbook.add_format({
                'font_size': 8 if level >= 5 else 9,
                'bg_color': bg_color,
                'font_color': zero_color,
                'border': 1,
                'num_format': '#,##0.00',
                'align': 'right',
            })

        # Total con colores
        total_bg = '#2c3e50'
        formats['total_positive'] = workbook.add_format({
            'bold': True,
            'font_size': 10,
            'bg_color': total_bg,
            'font_color': '#90EE90',  # Verde claro para fondo oscuro
            'border': 2,
            'num_format': '#,##0.00',
            'align': 'right',
        })
        formats['total_negative'] = workbook.add_format({
            'bold': True,
            'font_size': 10,
            'bg_color': total_bg,
            'font_color': '#FF6B6B',  # Rojo claro para fondo oscuro
            'border': 2,
            'num_format': '(#,##0.00)',
            'align': 'right',
        })
        formats['total_zero'] = workbook.add_format({
            'bold': True,
            'font_size': 10,
            'bg_color': total_bg,
            'font_color': '#CCCCCC',  # Gris claro
            'border': 2,
            'num_format': '#,##0.00',
            'align': 'right',
        })

        return formats

    def _write_xlsx_header(self, sheet, formats, options):
        """Escribe el encabezado del reporte."""
        company = self.env.company

        # Título
        sheet.merge_range('A1:I1', company.name, formats['title'])
        sheet.merge_range('A2:I2', f"NIT: {company.vat or ''}", formats['subtitle'])
        sheet.merge_range('A3:I3', 'BALANCE DE PRUEBA POR TERCERO', formats['title'])

        # Fechas
        date_from = options.get('date', {}).get('date_from', '')
        date_to = options.get('date', {}).get('date_to', '')
        sheet.merge_range('A4:I4', f"Del {date_from} al {date_to}", formats['subtitle'])

        return 6  # Siguiente fila disponible

    def _write_xlsx_column_headers(self, sheet, formats, row):
        """Escribe los encabezados de columnas."""
        headers = [
            ('CÓDIGO', 10),
            ('NOMBRE CUENTA', 35),
            ('TERCERO', 30),
            ('NIT/CC', 15),
            ('ÚLT.MOV', 12),
            ('SALDO INICIAL', 15),
            ('DÉBITOS', 15),
            ('CRÉDITOS', 15),
            ('SALDO FINAL', 15),
        ]

        for col, (header, width) in enumerate(headers):
            if col >= 5:  # Columnas numéricas
                sheet.write(row, col, header, formats['header_right'])
            else:
                sheet.write(row, col, header, formats['header'])

        return row + 1

    def _write_xlsx_data_with_outlines(self, sheet, formats, lines, start_row, options):
        """
        Escribe los datos con outlines/agrupamiento expandido.

        Según la documentación de XlsxWriter:
        - set_row con 'level' agrupa filas adyacentes con el mismo nivel
        - Los símbolos +/- aparecen automáticamente cuando hay filas agrupadas
        - Con outline_settings(False, False, ...) los símbolos aparecen ARRIBA del grupo

        Estructura PUC:
        - Nivel 1 (Clase): level=0 (sin agrupar, siempre visible)
        - Nivel 2 (Grupo): level=1
        - Nivel 3 (Cuenta): level=2
        - Nivel 4 (Subcuenta): level=3
        - Nivel 5 (Tercero): level=4

        Las filas están VISIBLES (no hidden) para que el usuario pueda colapsar manualmente.
        """
        rows_data = []

        def get_col_value(line, label):
            """Obtiene el valor de una columna por su expression_label."""
            for col in line.get('columns', []):
                if col.get('expression_label') == label:
                    return col.get('no_format', '')
            return ''

        def to_float(value):
            """Convierte a float de forma segura."""
            try:
                return float(value) if value else 0
            except (ValueError, TypeError):
                return 0

        for line in lines:
            level = line.get('level', 1)
            is_total = 'total' in str(line.get('id', '')).lower()

            rows_data.append({
                'level': level,
                'is_total': is_total,
                'account_code': get_col_value(line, 'account_code') or '',
                'account_name': get_col_value(line, 'account_name') or '' if level < 5 else '',
                'partner_name': get_col_value(line, 'partner_name') or '',
                'partner_vat': get_col_value(line, 'partner_vat') or '',
                'last_move_date': get_col_value(line, 'last_move_date'),
                'initial_balance': to_float(get_col_value(line, 'initial_balance')),
                'debit': to_float(get_col_value(line, 'debit')),
                'credit': to_float(get_col_value(line, 'credit')),
                'final_balance': to_float(get_col_value(line, 'final_balance')),
            })

        # Escribir datos con outlines
        row = start_row
        for i, data in enumerate(rows_data):
            level = data['level']
            is_total = data['is_total']

            # Determinar formato según nivel
            if is_total:
                fmt = 'total'
                fmt_num = 'total_num'
            else:
                fmt = f'level_{level}'
                fmt_num = f'level_{level}_num'

            # Configurar outline level para cada fila
            # Nivel 1 no tiene outline (level=0 o sin configurar)
            # Niveles 2-5 tienen outline levels 1-4
            if not is_total and level >= 2:
                outline_level = level - 1  # level 2 -> outline 1, etc.
                sheet.set_row(row, None, None, {'level': outline_level})

            # Escribir datos de la fila
            text_format = formats.get(fmt, formats['level_5'])
            num_format = formats.get(fmt_num, formats['level_5_num'])

            # Verificar si usar colores
            use_colors = options.get('use_color_formatting', False)

            sheet.write(row, 0, data['account_code'], text_format)
            sheet.write(row, 1, data['account_name'], text_format)
            sheet.write(row, 2, data['partner_name'], text_format)
            sheet.write(row, 3, data['partner_vat'], text_format)
            sheet.write(row, 4, str(data['last_move_date']) if data['last_move_date'] else '', text_format)

            # Función para obtener formato coloreado
            def get_colored_format(value, base_level, is_total_line):
                if not use_colors:
                    return num_format
                if is_total_line:
                    if value > 0:
                        return formats.get('total_positive', num_format)
                    elif value < 0:
                        return formats.get('total_negative', num_format)
                    else:
                        return formats.get('total_zero', num_format)
                else:
                    lvl = min(max(base_level, 4), 6)
                    if value > 0:
                        return formats.get(f'level_{lvl}_positive', num_format)
                    elif value < 0:
                        return formats.get(f'level_{lvl}_negative', num_format)
                    else:
                        return formats.get(f'level_{lvl}_zero', num_format)

            # Escribir números con formato de color si está habilitado
            sheet.write_number(row, 5, data['initial_balance'], get_colored_format(data['initial_balance'], level, is_total))
            sheet.write_number(row, 6, data['debit'], get_colored_format(data['debit'], level, is_total))
            sheet.write_number(row, 7, data['credit'], get_colored_format(data['credit'], level, is_total))
            sheet.write_number(row, 8, data['final_balance'], get_colored_format(data['final_balance'], level, is_total))

            row += 1

        return row

    def _set_xlsx_column_widths(self, sheet):
        """Configura los anchos de columna."""
        widths = [10, 35, 30, 15, 12, 15, 15, 15, 15]
        for col, width in enumerate(widths):
            sheet.set_column(col, col, width)

    def _export_trial_balance_partner_pdf(self, options):
        """
        Genera PDF para Balance de Prueba por Tercero con HTML autocontenido.
        Incluye pie de página con número de página.
        """
        print_options = self.get_options(previous_options={**options, 'export_mode': 'print'})

        # Obtener líneas expandidas - usar el modelo del handler directamente
        handler_model = self.custom_handler_model_id.model
        handler = self.env[handler_model]
        lines = handler._get_lines_for_pdf(self, print_options)
        lines = self._format_lines_for_display(lines, print_options)

        # Generar HTML autocontenido (cuerpo)
        html_content = self._get_trial_balance_partner_pdf_html(print_options, lines)

        # Generar HTML del pie de página
        footer_html = self._get_trial_balance_partner_footer_html(print_options)

        # Obtener paperformat
        paperformat = self.env.ref('libros_contables_colombia.paperformat_trial_balance_partner', raise_if_not_found=False)
        if not paperformat:
            paperformat = self.env['report.paperformat'].search([('format', '=', 'Letter')], limit=1)

        # Configuración de wkhtmltopdf
        specific_paperformat_args = {
            'data-report-margin-top': 5,
            'data-report-header-spacing': 2,
            'data-report-margin-bottom': 12,
        }

        # Generar PDF con footer que incluye número de página
        action_report = self.env['ir.actions.report']
        pdf_content = action_report.with_context(
            force_report_rendering=True,
            paperformat_id=paperformat.id if paperformat else False
        )._run_wkhtmltopdf(
            [html_content],
            footer=footer_html,
            landscape=True,
            specific_paperformat_args=specific_paperformat_args
        )

        return {
            'file_name': self.get_default_report_filename(options, 'pdf'),
            'file_content': pdf_content,
            'file_type': 'pdf',
        }

    def _get_pdf_export_html(self, options, lines, additional_context=None, template=None):
        """
        Sobrescribe para usar template personalizado en Balance por Tercero.
        """
        report_info = self.get_report_information(options)

        custom_print_templates = report_info['custom_display'].get('pdf_export', {})
        template = custom_print_templates.get('pdf_export_main', 'account_reports.pdf_export_main')

        render_values = {
            'report': self,
            'report_title': self.name,
            'options': options,
            'table_start': markupsafe.Markup('<tbody>'),
            'table_end': markupsafe.Markup('''
                </tbody></table></div>
                <div style="page-break-after: always"></div>
                <div class="d-flex align-items-start">
                <table class="o_table">
            '''),
            'column_headers_render_data': self._get_column_headers_render_data(options),
            'custom_templates': custom_print_templates,
        }
        if additional_context:
            render_values.update(additional_context)

        if options.get('order_column'):
            lines = self.sort_lines(lines, options)

        render_values['lines'] = lines

        # Manage annotations.
        render_values['annotations'] = self._build_annotations_list_for_pdf_export(
            options['date'], lines, report_info['annotations']
        )

        options['css_custom_class'] = report_info['custom_display'].get('css_custom_class', '')

        # Render.
        return self.env['ir.qweb']._render(template, render_values)

    # =========================================================================
    # MÉTODOS PARA BALANCE PUC - PDF Y EXCEL
    # =========================================================================

    def _export_trial_balance_puc_pdf(self, options):
        """
        Genera PDF para Balance de Prueba PUC con HTML autocontenido.
        """
        print_options = self.get_options(previous_options={**options, 'export_mode': 'print'})

        # Obtener líneas expandidas desde el handler
        handler_model = self.custom_handler_model_id.model
        handler = self.env[handler_model]
        lines = handler._get_lines_for_pdf(self, print_options)

        # Generar HTML autocontenido
        html_content = self._get_trial_balance_puc_pdf_html(print_options, lines)

        # Generar HTML del pie de página
        footer_html = self._get_trial_balance_puc_footer_html(print_options)

        # Obtener paperformat
        paperformat = self.env.ref('libros_contables_colombia.paperformat_trial_balance_partner', raise_if_not_found=False)
        if not paperformat:
            paperformat = self.env['report.paperformat'].search([('format', '=', 'Letter')], limit=1)

        # Configuración de wkhtmltopdf
        specific_paperformat_args = {
            'data-report-margin-top': 5,
            'data-report-header-spacing': 2,
            'data-report-margin-bottom': 12,
        }

        # Generar PDF
        action_report = self.env['ir.actions.report']
        pdf_content = action_report.with_context(
            force_report_rendering=True,
            paperformat_id=paperformat.id if paperformat else False
        )._run_wkhtmltopdf(
            [html_content],
            footer=footer_html,
            landscape=True,
            specific_paperformat_args=specific_paperformat_args
        )

        return {
            'file_name': self.get_default_report_filename(options, 'pdf'),
            'file_content': pdf_content,
            'file_type': 'pdf',
        }

    def _get_trial_balance_puc_pdf_html(self, options, lines):
        """
        Genera HTML autocontenido para el Balance de Prueba PUC.
        """
        company = self.env.company

        # Logo como base64 data URI
        logo_html = ''
        if company.logo:
            logo_b64 = base64.b64encode(company.logo).decode('utf-8')
            logo_html = f'<img src="data:image/png;base64,{logo_b64}" style="max-height: 45px; max-width: 90px;" alt="Logo"/>'

        # Fechas
        date_from = options.get('date', {}).get('date_from', '')
        date_to = options.get('date', {}).get('date_to', '')

        # Nombre de empresa
        company_names = []
        for c in options.get('companies', []):
            company_names.append(c.get('name', ''))
        company_name = ', '.join(company_names) if company_names else company.name

        def format_money(val):
            """Formatea valor monetario."""
            if val is None:
                return ''
            try:
                return f"$ {float(val):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            except (ValueError, TypeError):
                return ''

        def format_date(val):
            """Formatea fecha."""
            if val is None:
                return ''
            from datetime import date as dt_date
            if isinstance(val, dt_date):
                return val.strftime('%d/%m/%Y')
            return str(val) if val else ''

        # Generar filas de la tabla
        rows_html = []
        for line in lines:
            lvl = line.get('level', 1)
            is_total = line.get('is_total', False)
            code = line.get('code', '')
            name = line.get('name', '')

            # Estilos según nivel
            if is_total:
                style = 'background: #2c3e50; color: white; font-weight: bold;'
            elif lvl == 1:
                style = 'background: #34495e; color: white; font-weight: bold;'
            elif lvl == 2:
                style = 'background: #7f8c8d; color: white; font-weight: bold;'
            elif lvl == 3:
                style = 'background: #bdc3c7; color: #2c3e50; font-weight: bold;'
            elif lvl == 4:
                style = 'background: #ecf0f1; color: #2c3e50;'
            else:
                style = 'background: #fafafa; color: #333333; font-style: italic;'

            # Indentación según nivel
            indent = (lvl - 1) * 15

            row_html = f'''
            <tr style="{style}">
                <td style="padding: 2px 4px; padding-left: {indent + 4}px;">{code}</td>
                <td style="padding: 2px 4px; padding-left: {indent + 4}px;">{name}</td>
                <td style="padding: 2px 4px; text-align: right;">{format_money(line.get('initial', 0))}</td>
                <td style="padding: 2px 4px; text-align: right;">{format_money(line.get('debit', 0))}</td>
                <td style="padding: 2px 4px; text-align: right;">{format_money(line.get('credit', 0))}</td>
                <td style="padding: 2px 4px; text-align: right;">{format_money(line.get('final', 0))}</td>
                <td style="padding: 2px 4px; text-align: center;">{format_date(line.get('last_move_date'))}</td>
            </tr>
            '''
            rows_html.append(row_html)

        return f'''<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <style>
        @page {{
            size: letter landscape;
            margin: 5mm 10mm 15mm 10mm;
        }}
        body {{
            font-family: Arial, Helvetica, sans-serif;
            font-size: 8pt;
            margin: 0;
            padding: 0;
        }}
        .header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 10px;
            border-bottom: 2px solid #2c3e50;
            padding-bottom: 5px;
        }}
        .company-info {{
            text-align: center;
            flex-grow: 1;
        }}
        .company-name {{
            font-size: 12pt;
            font-weight: bold;
        }}
        .report-title {{
            font-size: 11pt;
            font-weight: bold;
            margin: 3px 0;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 7pt;
        }}
        th {{
            background: #2c3e50;
            color: white;
            padding: 4px;
            text-align: center;
            font-weight: bold;
        }}
        td {{
            border: 1px solid #ddd;
            padding: 2px 4px;
        }}
    </style>
</head>
<body>
    <div class="header">
        <div>{logo_html}</div>
        <div class="company-info">
            <div class="company-name">{company_name}</div>
            <div>NIT: {company.vat or ''}</div>
            <div class="report-title">BALANCE DE PRUEBA PUC</div>
            <div>Del {date_from} al {date_to}</div>
        </div>
        <div style="width: 90px;"></div>
    </div>
    <table>
        <thead>
            <tr>
                <th style="width: 80px;">CÓDIGO</th>
                <th style="width: 250px;">NOMBRE CUENTA</th>
                <th style="width: 100px;">SALDO INICIAL</th>
                <th style="width: 100px;">DÉBITOS</th>
                <th style="width: 100px;">CRÉDITOS</th>
                <th style="width: 100px;">SALDO FINAL</th>
                <th style="width: 70px;">ÚLT. MOV.</th>
            </tr>
        </thead>
        <tbody>
            {''.join(rows_html)}
        </tbody>
    </table>
</body>
</html>'''

    def _get_trial_balance_puc_footer_html(self, options):
        """Genera HTML para el pie de página del PDF del Balance PUC."""
        return '''<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <style>
        body {
            font-family: Arial, sans-serif;
            font-size: 8pt;
            margin: 0;
            padding: 0 10px;
        }
        table {
            width: 100%;
            border-collapse: collapse;
        }
        td {
            padding: 2px 5px;
        }
        .line {
            border-top: 1px solid #ccc;
        }
    </style>
</head>
<body>
    <table>
        <tr class="line">
            <td style="text-align: left;">Balance de Prueba PUC</td>
            <td style="text-align: center;">Generado: ''' + fields.Datetime.now().strftime('%d/%m/%Y %H:%M') + '''</td>
            <td style="text-align: right;">
                Página <span id="page"></span> de <span id="topage"></span>
            </td>
        </tr>
    </table>
</body>
</html>'''

    def _export_trial_balance_puc_xlsx(self, options):
        """
        Genera Excel para Balance de Prueba PUC con agrupamiento por niveles.
        Usa outlines de Excel para permitir expandir/contraer niveles.
        """
        if not xlsxwriter:
            raise ValueError("xlsxwriter no está instalado")

        print_options = self.get_options(previous_options={**options, 'export_mode': 'print'})

        # Obtener líneas expandidas
        handler_model = self.custom_handler_model_id.model
        handler = self.env[handler_model]
        lines = handler._get_lines_for_pdf(self, print_options)

        # Crear workbook
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {
            'in_memory': True,
            'strings_to_formulas': False,
        })

        # Configurar outline
        sheet = workbook.add_worksheet('Balance PUC')
        sheet.outline_settings(True, False, False, False)

        # Definir formatos
        formats = self._get_xlsx_formats_puc(workbook)

        # Escribir encabezado
        row = self._write_xlsx_header_puc(sheet, formats, print_options)

        # Escribir encabezados de columnas
        row = self._write_xlsx_column_headers_puc(sheet, formats, row)

        # Escribir datos con outlines (con soporte de colores)
        row = self._write_xlsx_data_with_outlines_puc(sheet, formats, lines, row, print_options)

        # Ajustar anchos de columna
        self._set_xlsx_column_widths_puc(sheet)

        workbook.close()
        output.seek(0)

        # Post-procesar con openpyxl para asegurar showOutlineSymbols=True
        if openpyxl_load_workbook:
            try:
                wb = openpyxl_load_workbook(output)
                ws = wb.active
                for view in ws.views.sheetView:
                    view.showOutlineSymbols = True

                final_output = io.BytesIO()
                wb.save(final_output)
                final_output.seek(0)
                generated_file = final_output.read()
                final_output.close()
            except Exception as e:
                _logger.warning("Error al post-procesar Excel con openpyxl: %s", str(e))
                output.seek(0)
                generated_file = output.read()
        else:
            generated_file = output.read()

        output.close()

        return {
            'file_name': self.get_default_report_filename(options, 'xlsx'),
            'file_content': generated_file,
            'file_type': 'xlsx',
        }

    def _get_xlsx_formats_puc(self, workbook):
        """Define los formatos para el Excel del Balance PUC."""
        formats = {}

        # Formato de título
        formats['title'] = workbook.add_format({
            'bold': True,
            'font_size': 14,
            'align': 'center',
            'valign': 'vcenter',
        })

        # Formato de subtítulo
        formats['subtitle'] = workbook.add_format({
            'font_size': 10,
            'align': 'center',
            'valign': 'vcenter',
        })

        # Formato de encabezado de columna
        formats['header'] = workbook.add_format({
            'bold': True,
            'font_size': 9,
            'bg_color': '#2c3e50',
            'font_color': 'white',
            'align': 'center',
            'valign': 'vcenter',
            'border': 1,
            'text_wrap': True,
        })

        formats['header_right'] = workbook.add_format({
            'bold': True,
            'font_size': 9,
            'bg_color': '#2c3e50',
            'font_color': 'white',
            'align': 'right',
            'valign': 'vcenter',
            'border': 1,
        })

        # Nivel 1 - Clase
        formats['level_1'] = workbook.add_format({
            'bold': True,
            'font_size': 10,
            'bg_color': '#34495e',
            'font_color': 'white',
            'border': 1,
        })
        formats['level_1_num'] = workbook.add_format({
            'bold': True,
            'font_size': 10,
            'bg_color': '#34495e',
            'font_color': 'white',
            'border': 1,
            'num_format': '#,##0.00',
            'align': 'right',
        })
        formats['level_1_date'] = workbook.add_format({
            'bold': True,
            'font_size': 10,
            'bg_color': '#34495e',
            'font_color': 'white',
            'border': 1,
            'align': 'center',
            'num_format': 'dd/mm/yyyy',
        })

        # Nivel 2 - Grupo
        formats['level_2'] = workbook.add_format({
            'bold': True,
            'font_size': 9,
            'bg_color': '#7f8c8d',
            'font_color': 'white',
            'border': 1,
        })
        formats['level_2_num'] = workbook.add_format({
            'bold': True,
            'font_size': 9,
            'bg_color': '#7f8c8d',
            'font_color': 'white',
            'border': 1,
            'num_format': '#,##0.00',
            'align': 'right',
        })
        formats['level_2_date'] = workbook.add_format({
            'bold': True,
            'font_size': 9,
            'bg_color': '#7f8c8d',
            'font_color': 'white',
            'border': 1,
            'align': 'center',
            'num_format': 'dd/mm/yyyy',
        })

        # Nivel 3 - Cuenta
        formats['level_3'] = workbook.add_format({
            'bold': True,
            'font_size': 9,
            'bg_color': '#bdc3c7',
            'font_color': '#2c3e50',
            'border': 1,
        })
        formats['level_3_num'] = workbook.add_format({
            'bold': True,
            'font_size': 9,
            'bg_color': '#bdc3c7',
            'font_color': '#2c3e50',
            'border': 1,
            'num_format': '#,##0.00',
            'align': 'right',
        })
        formats['level_3_date'] = workbook.add_format({
            'bold': True,
            'font_size': 9,
            'bg_color': '#bdc3c7',
            'font_color': '#2c3e50',
            'border': 1,
            'align': 'center',
            'num_format': 'dd/mm/yyyy',
        })

        # Nivel 4 - Subcuenta
        formats['level_4'] = workbook.add_format({
            'font_size': 9,
            'bg_color': '#ecf0f1',
            'font_color': '#2c3e50',
            'border': 1,
        })
        formats['level_4_num'] = workbook.add_format({
            'font_size': 9,
            'bg_color': '#ecf0f1',
            'font_color': '#2c3e50',
            'border': 1,
            'num_format': '#,##0.00',
            'align': 'right',
        })
        formats['level_4_date'] = workbook.add_format({
            'font_size': 9,
            'bg_color': '#ecf0f1',
            'font_color': '#2c3e50',
            'border': 1,
            'align': 'center',
            'num_format': 'dd/mm/yyyy',
        })

        # Nivel 5 - Auxiliar
        formats['level_5'] = workbook.add_format({
            'font_size': 8,
            'bg_color': '#fafafa',
            'font_color': '#333333',
            'border': 1,
            'italic': True,
        })
        formats['level_5_num'] = workbook.add_format({
            'font_size': 8,
            'bg_color': '#fafafa',
            'font_color': '#333333',
            'border': 1,
            'num_format': '#,##0.00',
            'align': 'right',
        })
        formats['level_5_date'] = workbook.add_format({
            'font_size': 8,
            'bg_color': '#fafafa',
            'font_color': '#333333',
            'border': 1,
            'align': 'center',
            'num_format': 'dd/mm/yyyy',
        })

        # Total general
        formats['total'] = workbook.add_format({
            'bold': True,
            'font_size': 10,
            'bg_color': '#2c3e50',
            'font_color': 'white',
            'border': 2,
        })
        formats['total_num'] = workbook.add_format({
            'bold': True,
            'font_size': 10,
            'bg_color': '#2c3e50',
            'font_color': 'white',
            'border': 2,
            'num_format': '#,##0.00',
            'align': 'right',
        })
        formats['total_date'] = workbook.add_format({
            'bold': True,
            'font_size': 10,
            'bg_color': '#2c3e50',
            'font_color': 'white',
            'border': 2,
            'align': 'center',
            'num_format': 'dd/mm/yyyy',
        })

        # =========================================================================
        # FORMATOS CON COLORES (para use_color_formatting)
        # =========================================================================
        positive_color = '#198754'  # Verde Bootstrap
        negative_color = '#dc3545'  # Rojo Bootstrap
        zero_color = '#999999'      # Gris

        # Niveles con colores (4-6 son los de detalle)
        for level in [4, 5, 6]:
            bg_color = '#fafafa' if level == 5 else '#ecf0f1' if level == 4 else '#ffffff'

            formats[f'level_{level}_positive'] = workbook.add_format({
                'font_size': 8 if level >= 5 else 9,
                'bg_color': bg_color,
                'font_color': positive_color,
                'border': 1,
                'num_format': '#,##0.00',
                'align': 'right',
            })
            formats[f'level_{level}_negative'] = workbook.add_format({
                'font_size': 8 if level >= 5 else 9,
                'bg_color': bg_color,
                'font_color': negative_color,
                'border': 1,
                'num_format': '(#,##0.00)',
                'align': 'right',
            })
            formats[f'level_{level}_zero'] = workbook.add_format({
                'font_size': 8 if level >= 5 else 9,
                'bg_color': bg_color,
                'font_color': zero_color,
                'border': 1,
                'num_format': '#,##0.00',
                'align': 'right',
            })

        # Total con colores
        total_bg = '#2c3e50'
        formats['total_positive'] = workbook.add_format({
            'bold': True,
            'font_size': 10,
            'bg_color': total_bg,
            'font_color': '#90EE90',
            'border': 2,
            'num_format': '#,##0.00',
            'align': 'right',
        })
        formats['total_negative'] = workbook.add_format({
            'bold': True,
            'font_size': 10,
            'bg_color': total_bg,
            'font_color': '#FF6B6B',
            'border': 2,
            'num_format': '(#,##0.00)',
            'align': 'right',
        })
        formats['total_zero'] = workbook.add_format({
            'bold': True,
            'font_size': 10,
            'bg_color': total_bg,
            'font_color': '#CCCCCC',
            'border': 2,
            'num_format': '#,##0.00',
            'align': 'right',
        })

        return formats

    def _write_xlsx_header_puc(self, sheet, formats, options):
        """Escribe el encabezado del reporte Balance PUC."""
        company = self.env.company

        # Título
        sheet.merge_range('A1:G1', company.name, formats['title'])
        sheet.merge_range('A2:G2', f"NIT: {company.vat or ''}", formats['subtitle'])
        sheet.merge_range('A3:G3', 'BALANCE DE PRUEBA PUC', formats['title'])

        # Fechas
        date_from = options.get('date', {}).get('date_from', '')
        date_to = options.get('date', {}).get('date_to', '')
        sheet.merge_range('A4:G4', f"Del {date_from} al {date_to}", formats['subtitle'])

        return 6  # Siguiente fila disponible

    def _write_xlsx_column_headers_puc(self, sheet, formats, row):
        """Escribe los encabezados de columnas para Balance PUC."""
        headers = [
            ('CÓDIGO', 12),
            ('NOMBRE CUENTA', 45),
            ('SALDO INICIAL', 15),
            ('DÉBITOS', 15),
            ('CRÉDITOS', 15),
            ('SALDO FINAL', 15),
            ('ÚLT. MOV.', 12),
        ]

        for col, (header, width) in enumerate(headers):
            if col >= 2 and col <= 5:  # Columnas numéricas
                sheet.write(row, col, header, formats['header_right'])
            else:
                sheet.write(row, col, header, formats['header'])

        return row + 1

    def _write_xlsx_data_with_outlines_puc(self, sheet, formats, lines, start_row, options=None):
        """
        Escribe los datos del Balance PUC con outlines.
        Soporta formato de colores cuando use_color_formatting está habilitado.
        """
        row = start_row
        options = options or {}
        use_colors = options.get('use_color_formatting', False)

        # Función para obtener formato coloreado
        def get_colored_format(value, base_level, is_total_line, default_format):
            if not use_colors:
                return default_format
            if is_total_line:
                if value > 0:
                    return formats.get('total_positive', default_format)
                elif value < 0:
                    return formats.get('total_negative', default_format)
                else:
                    return formats.get('total_zero', default_format)
            else:
                lvl = min(max(base_level, 4), 6)
                if value > 0:
                    return formats.get(f'level_{lvl}_positive', default_format)
                elif value < 0:
                    return formats.get(f'level_{lvl}_negative', default_format)
                else:
                    return formats.get(f'level_{lvl}_zero', default_format)

        for line in lines:
            level = line.get('level', 1)
            is_total = line.get('is_total', False)

            # Determinar formato según nivel
            if is_total:
                fmt = 'total'
                fmt_num = 'total_num'
                fmt_date = 'total_date'
            else:
                fmt = f'level_{level}'
                fmt_num = f'level_{level}_num'
                fmt_date = f'level_{level}_date'

            # Configurar outline level
            if not is_total and level >= 2:
                outline_level = level - 1
                sheet.set_row(row, None, None, {'level': outline_level})

            # Escribir datos de la fila
            text_format = formats.get(fmt, formats['level_5'])
            num_format = formats.get(fmt_num, formats['level_5_num'])
            date_format = formats.get(fmt_date, formats['level_5_date'])

            sheet.write(row, 0, line.get('code', ''), text_format)
            sheet.write(row, 1, line.get('name', ''), text_format)

            # Obtener valores numéricos
            initial_val = float(line.get('initial', 0) or 0)
            debit_val = float(line.get('debit', 0) or 0)
            credit_val = float(line.get('credit', 0) or 0)
            final_val = float(line.get('final', 0) or 0)

            # Escribir números con formato de color
            sheet.write_number(row, 2, initial_val, get_colored_format(initial_val, level, is_total, num_format))
            sheet.write_number(row, 3, debit_val, get_colored_format(debit_val, level, is_total, num_format))
            sheet.write_number(row, 4, credit_val, get_colored_format(credit_val, level, is_total, num_format))
            sheet.write_number(row, 5, final_val, get_colored_format(final_val, level, is_total, num_format))

            # Fecha
            last_move = line.get('last_move_date')
            if last_move:
                sheet.write(row, 6, last_move, date_format)
            else:
                sheet.write(row, 6, '', text_format)

            row += 1

        return row

    def _set_xlsx_column_widths_puc(self, sheet):
        """Configura los anchos de columna para Balance PUC."""
        widths = [12, 45, 15, 15, 15, 15, 12]
        for col, width in enumerate(widths):
            sheet.set_column(col, col, width)

    # =========================================================================
    # MÉTODOS PARA LIBRO AUXILIAR - PDF Y EXCEL
    # =========================================================================

    def _export_auxiliary_book_pdf(self, options):
        """
        Genera PDF para Libro Auxiliar con HTML autocontenido.
        Incluye información de auditoría y filtros aplicados.
        """
        print_options = self.get_options(previous_options={**options, 'export_mode': 'print'})

        # Obtener líneas expandidas desde el handler
        handler_model = self.custom_handler_model_id.model
        handler = self.env[handler_model]
        lines = handler._get_lines_for_pdf(self, print_options)

        # Determinar si es analítico
        is_analytic = self._is_auxiliary_book_analytic(options)

        # Generar HTML autocontenido
        html_content = self._get_auxiliary_book_pdf_html(print_options, lines, is_analytic)

        # Generar HTML del pie de página
        footer_html = self._get_auxiliary_book_footer_html(print_options, is_analytic)

        # Obtener paperformat
        paperformat = self.env.ref('libros_contables_colombia.paperformat_trial_balance_partner', raise_if_not_found=False)
        if not paperformat:
            paperformat = self.env['report.paperformat'].search([('format', '=', 'Letter')], limit=1)

        # Configuración de wkhtmltopdf
        specific_paperformat_args = {
            'data-report-margin-top': 5,
            'data-report-header-spacing': 2,
            'data-report-margin-bottom': 12,
        }

        # Generar PDF
        action_report = self.env['ir.actions.report']
        pdf_content = action_report.with_context(
            force_report_rendering=True,
            paperformat_id=paperformat.id if paperformat else False
        )._run_wkhtmltopdf(
            [html_content],
            footer=footer_html,
            landscape=True,
            specific_paperformat_args=specific_paperformat_args
        )

        return {
            'file_name': self.get_default_report_filename(options, 'pdf'),
            'file_content': pdf_content,
            'file_type': 'pdf',
        }

    def _get_auxiliary_book_pdf_html(self, options, lines, is_analytic=False):
        """
        Genera HTML autocontenido para el Libro Auxiliar.
        Optimizado con columnas relevantes y texto corto.
        Incluye información de filtros y auditoría.
        """
        company = self.env.company
        from datetime import datetime

        # Logo como base64 data URI
        logo_html = ''
        if company.logo:
            logo_b64 = base64.b64encode(company.logo).decode('utf-8')
            logo_html = f'<img src="data:image/png;base64,{logo_b64}" style="max-height: 40px; max-width: 80px;" alt="Logo"/>'

        # Fechas y datos de auditoría
        date_from = options.get('date', {}).get('date_from', '')
        date_to = options.get('date', {}).get('date_to', '')
        generated_at = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        user_name = self.env.user.name

        # Nombre de empresa
        company_names = []
        for c in options.get('companies', []):
            company_names.append(c.get('name', ''))
        company_name = ', '.join(company_names) if company_names else company.name

        # Filtros aplicados
        filters_info = []
        if options.get('journals'):
            journal_names = [j.get('name', '') for j in options.get('journals', []) if j.get('selected')]
            if journal_names:
                filters_info.append(f"Diarios: {', '.join(journal_names[:3])}{'...' if len(journal_names) > 3 else ''}")
        if options.get('partner_ids'):
            filters_info.append(f"Terceros: {len(options.get('partner_ids', []))} seleccionados")

        filters_html = f'<div style="font-size: 6pt; color: #666;">Filtros: {" | ".join(filters_info) if filters_info else "Ninguno"}</div>'

        # Título del reporte
        report_title = 'LIBRO AUXILIAR ANALÍTICO' if is_analytic else 'LIBRO AUXILIAR'

        def format_money(val):
            """Formatea valor monetario de forma compacta."""
            if val is None or val == 0:
                return ''
            try:
                v = float(val)
                if abs(v) >= 1000000:
                    return f"${v/1000000:,.1f}M".replace(',', 'X').replace('.', ',').replace('X', '.')
                return f"${v:,.0f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            except (ValueError, TypeError):
                return ''

        def format_date(val):
            """Formatea fecha de forma compacta."""
            if val is None:
                return ''
            from datetime import date as dt_date
            if isinstance(val, dt_date):
                return val.strftime('%d/%m')
            return str(val)[:5] if val else ''

        def truncate(text, max_len=20):
            """Trunca texto a longitud máxima."""
            if not text:
                return ''
            text = str(text)
            return text[:max_len] + '...' if len(text) > max_len else text

        # Generar filas de la tabla (columnas optimizadas para PDF)
        # Columnas: Cód|Cuenta|Fecha|Tipo|Doc|Etiq|Tercero|[Analít]|Base|Déb|Créd|Saldo
        rows_html = []
        for line in lines:
            lvl = line.get('level', 3)
            is_total = line.get('is_total', False)
            is_account = line.get('is_account', False)
            is_initial = line.get('is_initial', False)

            # Estilos según tipo de línea
            if is_total:
                style = 'background: #2c3e50; color: white; font-weight: bold; font-size: 7pt;'
            elif is_account:
                style = 'background: #34495e; color: white; font-weight: bold; font-size: 7pt;'
            elif is_initial:
                style = 'background: #ecf0f1; color: #2c3e50; font-style: italic; font-size: 6.5pt;'
            else:
                style = 'background: #fff; color: #333; font-size: 6.5pt;'

            # Construir fila
            row_cells = [
                f'<td style="padding: 1px 2px;">{truncate(line.get("account_code", ""), 8)}</td>',
                f'<td style="padding: 1px 2px;">{truncate(line.get("account_name", ""), 25)}</td>',
                f'<td style="padding: 1px 2px; text-align: center;">{format_date(line.get("date"))}</td>',
                f'<td style="padding: 1px 2px;">{truncate(line.get("journal_type", ""), 8)}</td>',
                f'<td style="padding: 1px 2px;">{truncate(line.get("move_name", ""), 12)}</td>',
                f'<td style="padding: 1px 2px;">{truncate(line.get("name", ""), 18)}</td>',
                f'<td style="padding: 1px 2px;">{truncate(line.get("partner_name", ""), 18)}</td>',
            ]

            # Agregar columna analítica si aplica
            if is_analytic:
                row_cells.append(f'<td style="padding: 1px 2px;">{truncate(line.get("analytic_account", ""), 15)}</td>')

            # Columnas numéricas
            row_cells.extend([
                f'<td style="padding: 1px 2px; text-align: right;">{format_money(line.get("tax_base_amount", 0))}</td>',
                f'<td style="padding: 1px 2px; text-align: right;">{format_money(line.get("debit", 0))}</td>',
                f'<td style="padding: 1px 2px; text-align: right;">{format_money(line.get("credit", 0))}</td>',
                f'<td style="padding: 1px 2px; text-align: right; font-weight: bold;">{format_money(line.get("balance", 0))}</td>',
            ])

            row_html = f'<tr style="{style}">{"".join(row_cells)}</tr>'
            rows_html.append(row_html)

        # Encabezados de columna
        headers = ['CÓD', 'CUENTA', 'FECHA', 'TIPO', 'DOC', 'ETIQ', 'TERCERO']
        if is_analytic:
            headers.append('ANALÍT')
        headers.extend(['BASE', 'DÉBITO', 'CRÉDITO', 'SALDO'])

        headers_html = ''.join([f'<th style="padding: 2px 3px; text-align: {"right" if h in ["BASE", "DÉBITO", "CRÉDITO", "SALDO"] else "left"};">{h}</th>' for h in headers])

        return f'''<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <style>
        @page {{
            size: letter landscape;
            margin: 5mm 8mm 12mm 8mm;
        }}
        body {{
            font-family: Arial, Helvetica, sans-serif;
            font-size: 7pt;
            margin: 0;
            padding: 0;
            -webkit-print-color-adjust: exact !important;
            print-color-adjust: exact !important;
        }}
        .header {{
            display: flex;
            justify-content: space-between;
            align-items: flex-start;
            margin-bottom: 5px;
            border-bottom: 2px solid #2c3e50;
            padding-bottom: 3px;
        }}
        .company-info {{
            text-align: center;
            flex-grow: 1;
        }}
        .company-name {{
            font-size: 10pt;
            font-weight: bold;
        }}
        .report-title {{
            font-size: 9pt;
            font-weight: bold;
            margin: 2px 0;
        }}
        .audit-info {{
            text-align: right;
            font-size: 6pt;
            color: #555;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 6.5pt;
            table-layout: fixed;
        }}
        th {{
            background: #2c3e50;
            color: white;
            padding: 2px 3px;
            font-weight: bold;
            font-size: 6.5pt;
            white-space: nowrap;
        }}
        td {{
            border-bottom: 1px solid #e0e0e0;
            white-space: nowrap;
            overflow: hidden;
            text-overflow: ellipsis;
        }}
    </style>
</head>
<body>
    <div class="header">
        <div style="width: 80px;">{logo_html}</div>
        <div class="company-info">
            <div class="company-name">{company_name}</div>
            <div style="font-size: 7pt;">NIT: {company.vat or ''}</div>
            <div class="report-title">{report_title}</div>
            <div style="font-size: 7pt;">Del {date_from} al {date_to}</div>
            {filters_html}
        </div>
        <div class="audit-info">
            <div><strong>Datos de Auditoría</strong></div>
            <div>Generado: {generated_at}</div>
            <div>Usuario: {user_name}</div>
            <div>Empresa: {company.name}</div>
        </div>
    </div>
    <table>
        <thead>
            <tr>{headers_html}</tr>
        </thead>
        <tbody>
            {''.join(rows_html)}
        </tbody>
    </table>
</body>
</html>'''

    def _get_auxiliary_book_footer_html(self, options, is_analytic=False):
        """Genera HTML para el pie de página del Libro Auxiliar."""
        report_name = 'Libro Auxiliar Analítico' if is_analytic else 'Libro Auxiliar'
        return f'''<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <script>
        function subst() {{
            var vars = {{}};
            var query_strings_from_url = document.location.search.substring(1).split('&');
            for (var query_string in query_strings_from_url) {{
                if (query_strings_from_url.hasOwnProperty(query_string)) {{
                    var temp_var = query_strings_from_url[query_string].split('=', 2);
                    vars[temp_var[0]] = decodeURIComponent(temp_var[1]);
                }}
            }}
            var page_elem = document.getElementById('page');
            var topage_elem = document.getElementById('topage');
            if (page_elem) page_elem.textContent = vars.page || '';
            if (topage_elem) topage_elem.textContent = vars.topage || '';
        }}
    </script>
    <style>
        body {{
            font-family: Arial, sans-serif;
            font-size: 7pt;
            margin: 0;
            padding: 0 10px;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
        }}
        td {{
            padding: 2px 5px;
        }}
        .line {{
            border-top: 1px solid #ccc;
        }}
    </style>
</head>
<body onload="subst()">
    <table>
        <tr class="line">
            <td style="text-align: left; color: #555;">{report_name} - Colombia</td>
            <td style="text-align: center; color: #555;">Generado: {fields.Datetime.now().strftime('%d/%m/%Y %H:%M')}</td>
            <td style="text-align: right; color: #333; font-weight: bold;">
                Página <span id="page"></span> de <span id="topage"></span>
            </td>
        </tr>
    </table>
</body>
</html>'''

    def _export_auxiliary_book_xlsx(self, options):
        """
        Genera Excel para Libro Auxiliar con filtros y outlines.
        """
        if not xlsxwriter:
            raise ValueError("xlsxwriter no está instalado")

        print_options = self.get_options(previous_options={**options, 'export_mode': 'print'})

        # Obtener líneas expandidas
        handler_model = self.custom_handler_model_id.model
        handler = self.env[handler_model]
        lines = handler._get_lines_for_pdf(self, print_options)

        # Determinar si es analítico
        is_analytic = self._is_auxiliary_book_analytic(options)

        # Crear workbook
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {
            'in_memory': True,
            'strings_to_formulas': False,
        })

        # Configurar outline
        sheet_name = 'Libro Auxiliar Analítico' if is_analytic else 'Libro Auxiliar'
        sheet = workbook.add_worksheet(sheet_name[:31])  # Excel limita a 31 caracteres
        sheet.outline_settings(True, False, False, False)

        # Definir formatos
        formats = self._get_xlsx_formats_auxiliary(workbook)

        # Escribir encabezado con info de auditoría
        row = self._write_xlsx_header_auxiliary(sheet, formats, print_options, is_analytic)

        # Escribir encabezados de columnas
        row = self._write_xlsx_column_headers_auxiliary(sheet, formats, row, is_analytic)

        # Escribir datos con outlines (con soporte de colores)
        row = self._write_xlsx_data_with_outlines_auxiliary(sheet, formats, lines, row, is_analytic, print_options)

        # Ajustar anchos de columna
        self._set_xlsx_column_widths_auxiliary(sheet, is_analytic)

        # Activar autofiltro en las columnas de datos
        filter_row = 7  # Fila donde están los encabezados
        num_cols = 16 if is_analytic else 15
        sheet.autofilter(filter_row, 0, row - 1, num_cols - 1)

        workbook.close()
        output.seek(0)

        # Post-procesar con openpyxl para asegurar showOutlineSymbols=True
        if openpyxl_load_workbook:
            try:
                wb = openpyxl_load_workbook(output)
                ws = wb.active
                for view in ws.views.sheetView:
                    view.showOutlineSymbols = True

                final_output = io.BytesIO()
                wb.save(final_output)
                final_output.seek(0)
                generated_file = final_output.read()
                final_output.close()
            except Exception as e:
                _logger.warning("Error al post-procesar Excel con openpyxl: %s", str(e))
                output.seek(0)
                generated_file = output.read()
        else:
            generated_file = output.read()

        output.close()

        return {
            'file_name': self.get_default_report_filename(options, 'xlsx'),
            'file_content': generated_file,
            'file_type': 'xlsx',
        }

    def _get_xlsx_formats_auxiliary(self, workbook):
        """Define los formatos para el Excel del Libro Auxiliar."""
        formats = {}

        # Formato de título
        formats['title'] = workbook.add_format({
            'bold': True,
            'font_size': 14,
            'align': 'center',
            'valign': 'vcenter',
        })

        # Formato de subtítulo
        formats['subtitle'] = workbook.add_format({
            'font_size': 10,
            'align': 'center',
            'valign': 'vcenter',
        })

        # Formato de info de auditoría
        formats['audit'] = workbook.add_format({
            'font_size': 8,
            'align': 'left',
            'font_color': '#666666',
        })

        # Formato de encabezado de columna
        formats['header'] = workbook.add_format({
            'bold': True,
            'font_size': 9,
            'bg_color': '#2c3e50',
            'font_color': 'white',
            'align': 'center',
            'valign': 'vcenter',
            'border': 1,
            'text_wrap': True,
        })

        formats['header_right'] = workbook.add_format({
            'bold': True,
            'font_size': 9,
            'bg_color': '#2c3e50',
            'font_color': 'white',
            'align': 'right',
            'valign': 'vcenter',
            'border': 1,
        })

        # Cuenta (nivel 1)
        formats['account'] = workbook.add_format({
            'bold': True,
            'font_size': 10,
            'bg_color': '#34495e',
            'font_color': 'white',
            'border': 1,
        })
        formats['account_num'] = workbook.add_format({
            'bold': True,
            'font_size': 10,
            'bg_color': '#34495e',
            'font_color': 'white',
            'border': 1,
            'num_format': '#,##0.00',
            'align': 'right',
        })
        formats['account_date'] = workbook.add_format({
            'bold': True,
            'font_size': 10,
            'bg_color': '#34495e',
            'font_color': 'white',
            'border': 1,
            'align': 'center',
            'num_format': 'dd/mm/yyyy',
        })

        # Saldo inicial (nivel 2)
        formats['initial'] = workbook.add_format({
            'font_size': 9,
            'bg_color': '#ecf0f1',
            'font_color': '#2c3e50',
            'border': 1,
            'italic': True,
        })
        formats['initial_num'] = workbook.add_format({
            'font_size': 9,
            'bg_color': '#ecf0f1',
            'font_color': '#2c3e50',
            'border': 1,
            'num_format': '#,##0.00',
            'align': 'right',
            'italic': True,
        })
        formats['initial_date'] = workbook.add_format({
            'font_size': 9,
            'bg_color': '#ecf0f1',
            'font_color': '#2c3e50',
            'border': 1,
            'align': 'center',
            'num_format': 'dd/mm/yyyy',
            'italic': True,
        })

        # Movimiento (nivel 3)
        formats['movement'] = workbook.add_format({
            'font_size': 9,
            'bg_color': '#ffffff',
            'font_color': '#333333',
            'border': 1,
        })
        formats['movement_num'] = workbook.add_format({
            'font_size': 9,
            'bg_color': '#ffffff',
            'font_color': '#333333',
            'border': 1,
            'num_format': '#,##0.00',
            'align': 'right',
        })
        formats['movement_date'] = workbook.add_format({
            'font_size': 9,
            'bg_color': '#ffffff',
            'font_color': '#333333',
            'border': 1,
            'align': 'center',
            'num_format': 'dd/mm/yyyy',
        })

        # Total general
        formats['total'] = workbook.add_format({
            'bold': True,
            'font_size': 10,
            'bg_color': '#2c3e50',
            'font_color': 'white',
            'border': 2,
        })
        formats['total_num'] = workbook.add_format({
            'bold': True,
            'font_size': 10,
            'bg_color': '#2c3e50',
            'font_color': 'white',
            'border': 2,
            'num_format': '#,##0.00',
            'align': 'right',
        })

        # =========================================================================
        # FORMATOS CON COLORES (para use_color_formatting)
        # =========================================================================
        positive_color = '#198754'  # Verde Bootstrap
        negative_color = '#dc3545'  # Rojo Bootstrap
        zero_color = '#999999'      # Gris

        # Movimiento con colores
        formats['movement_positive'] = workbook.add_format({
            'font_size': 9,
            'bg_color': '#ffffff',
            'font_color': positive_color,
            'border': 1,
            'num_format': '#,##0.00',
            'align': 'right',
        })
        formats['movement_negative'] = workbook.add_format({
            'font_size': 9,
            'bg_color': '#ffffff',
            'font_color': negative_color,
            'border': 1,
            'num_format': '(#,##0.00)',
            'align': 'right',
        })
        formats['movement_zero'] = workbook.add_format({
            'font_size': 9,
            'bg_color': '#ffffff',
            'font_color': zero_color,
            'border': 1,
            'num_format': '#,##0.00',
            'align': 'right',
        })

        # Total con colores
        total_bg = '#2c3e50'
        formats['total_positive'] = workbook.add_format({
            'bold': True,
            'font_size': 10,
            'bg_color': total_bg,
            'font_color': '#90EE90',  # Verde claro
            'border': 2,
            'num_format': '#,##0.00',
            'align': 'right',
        })
        formats['total_negative'] = workbook.add_format({
            'bold': True,
            'font_size': 10,
            'bg_color': total_bg,
            'font_color': '#FF6B6B',  # Rojo claro
            'border': 2,
            'num_format': '(#,##0.00)',
            'align': 'right',
        })
        formats['total_zero'] = workbook.add_format({
            'bold': True,
            'font_size': 10,
            'bg_color': total_bg,
            'font_color': '#CCCCCC',  # Gris claro
            'border': 2,
            'num_format': '#,##0.00',
            'align': 'right',
        })

        return formats

    def _write_xlsx_header_auxiliary(self, sheet, formats, options, is_analytic=False):
        """Escribe el encabezado del Libro Auxiliar con información de auditoría."""
        from datetime import datetime
        company = self.env.company
        report_title = 'LIBRO AUXILIAR ANALÍTICO' if is_analytic else 'LIBRO AUXILIAR'
        num_cols = 16 if is_analytic else 15

        # Título
        sheet.merge_range(0, 0, 0, num_cols - 1, company.name, formats['title'])
        sheet.merge_range(1, 0, 1, num_cols - 1, f"NIT: {company.vat or ''}", formats['subtitle'])
        sheet.merge_range(2, 0, 2, num_cols - 1, report_title, formats['title'])

        # Fechas
        date_from = options.get('date', {}).get('date_from', '')
        date_to = options.get('date', {}).get('date_to', '')
        sheet.merge_range(3, 0, 3, num_cols - 1, f"Del {date_from} al {date_to}", formats['subtitle'])

        # Información de auditoría
        generated_at = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        user_name = self.env.user.name
        sheet.write(5, 0, f"Generado: {generated_at} | Usuario: {user_name} | Empresa: {company.name}", formats['audit'])

        return 7  # Siguiente fila disponible

    def _write_xlsx_column_headers_auxiliary(self, sheet, formats, row, is_analytic=False):
        """Escribe los encabezados de columnas para Libro Auxiliar."""
        headers = [
            ('Código', 10),
            ('Cuenta', 30),
            ('Fecha', 10),
            ('Tipo Op.', 10),
            ('Secuencia', 10),
            ('Documento', 15),
            ('Etiqueta', 25),
            ('Ref.', 15),
            ('Ref. Pago', 15),
            ('Tercero', 25),
            ('Últ. Mov.', 10),
        ]

        if is_analytic:
            headers.append(('Cta. Analítica', 20))

        headers.extend([
            ('Base', 12),
            ('Débito', 15),
            ('Crédito', 15),
            ('Saldo', 15),
        ])

        for col, (header, width) in enumerate(headers):
            if header in ('Base', 'Débito', 'Crédito', 'Saldo'):
                sheet.write(row, col, header, formats['header_right'])
            else:
                sheet.write(row, col, header, formats['header'])

        return row + 1

    def _write_xlsx_data_with_outlines_auxiliary(self, sheet, formats, lines, start_row, is_analytic=False, options=None):
        """
        Escribe los datos del Libro Auxiliar con outlines.
        Soporta formato de colores cuando use_color_formatting está habilitado.
        """
        row = start_row
        options = options or {}
        use_colors = options.get('use_color_formatting', False)

        # Función para obtener formato coloreado para movimientos
        def get_colored_format(value, is_total_line, default_format):
            if not use_colors:
                return default_format
            if is_total_line:
                if value > 0:
                    return formats.get('total_positive', default_format)
                elif value < 0:
                    return formats.get('total_negative', default_format)
                else:
                    return formats.get('total_zero', default_format)
            else:
                if value > 0:
                    return formats.get('movement_positive', default_format)
                elif value < 0:
                    return formats.get('movement_negative', default_format)
                else:
                    return formats.get('movement_zero', default_format)

        for line in lines:
            is_total = line.get('is_total', False)
            is_account = line.get('is_account', False)
            is_initial = line.get('is_initial', False)

            # Determinar formato según tipo de línea
            if is_total:
                fmt = 'total'
                fmt_num = 'total_num'
                outline_level = None
            elif is_account:
                fmt = 'account'
                fmt_num = 'account_num'
                outline_level = None
            elif is_initial:
                fmt = 'initial'
                fmt_num = 'initial_num'
                outline_level = 1
            else:
                fmt = 'movement'
                fmt_num = 'movement_num'
                outline_level = 2

            # Configurar outline level
            if outline_level is not None:
                sheet.set_row(row, None, None, {'level': outline_level})

            # Obtener formatos
            text_format = formats.get(fmt, formats['movement'])
            num_format = formats.get(fmt_num, formats['movement_num'])
            date_format = formats.get(f'{fmt}_date', formats.get('movement_date', text_format))

            # Escribir columnas
            col = 0
            sheet.write(row, col, line.get('account_code', ''), text_format); col += 1
            sheet.write(row, col, line.get('account_name', ''), text_format); col += 1

            # Fecha
            date_val = line.get('date')
            if date_val:
                sheet.write(row, col, date_val, date_format)
            else:
                sheet.write(row, col, '', text_format)
            col += 1

            sheet.write(row, col, line.get('journal_type', ''), text_format); col += 1
            sheet.write(row, col, str(line.get('sequence', '') or ''), text_format); col += 1
            sheet.write(row, col, line.get('move_name', ''), text_format); col += 1
            sheet.write(row, col, line.get('name', ''), text_format); col += 1
            sheet.write(row, col, line.get('ref', ''), text_format); col += 1
            sheet.write(row, col, line.get('payment_ref', ''), text_format); col += 1
            sheet.write(row, col, line.get('partner_name', ''), text_format); col += 1

            # Última fecha de movimiento
            last_move = line.get('last_move_date')
            if last_move:
                sheet.write(row, col, last_move, date_format)
            else:
                sheet.write(row, col, '', text_format)
            col += 1

            # Columna analítica si aplica
            if is_analytic:
                sheet.write(row, col, line.get('analytic_account', ''), text_format)
                col += 1

            # Obtener valores numéricos
            tax_base_val = float(line.get('tax_base_amount', 0) or 0)
            debit_val = float(line.get('debit', 0) or 0)
            credit_val = float(line.get('credit', 0) or 0)
            balance_val = float(line.get('balance', 0) or 0)

            # Columnas numéricas con colores
            sheet.write_number(row, col, tax_base_val, get_colored_format(tax_base_val, is_total, num_format)); col += 1
            sheet.write_number(row, col, debit_val, get_colored_format(debit_val, is_total, num_format)); col += 1
            sheet.write_number(row, col, credit_val, get_colored_format(credit_val, is_total, num_format)); col += 1
            sheet.write_number(row, col, balance_val, get_colored_format(balance_val, is_total, num_format))

            row += 1

        return row

    def _set_xlsx_column_widths_auxiliary(self, sheet, is_analytic=False):
        """Configura los anchos de columna para Libro Auxiliar."""
        widths = [10, 30, 10, 10, 10, 15, 25, 15, 15, 25, 10]
        if is_analytic:
            widths.append(20)
        widths.extend([12, 15, 15, 15])

        for col, width in enumerate(widths):
            sheet.set_column(col, col, width)

    # =========================================================================
    # EXPORTACIÓN EXCEL CON COLORES Y ESTILOS PARA REPORTES COLOMBIANOS
    # =========================================================================

    def _inject_report_into_xlsx_sheet(self, options, workbook, sheet):
        """
        Override para agregar colores y estilos a la exportación Excel
        de reportes colombianos.

        Cuando use_color_formatting está habilitado:
        - Verde para valores positivos
        - Rojo para valores negativos
        - Símbolos contables (▲▼) si use_accounting_symbols está habilitado

        También aplica estilos de jerarquía con indentación visual.
        """
        # Verificar si es un reporte colombiano con formato de colores
        is_colombian = self._is_colombian_report()
        use_colors = options.get('use_color_formatting', False)
        use_symbols = options.get('use_accounting_symbols', False)
        color_mode = options.get('color_mode', 'accounting')

        # Si no es colombiano o no tiene colores, usar método original
        if not is_colombian or not use_colors:
            return super()._inject_report_into_xlsx_sheet(options, workbook, sheet)

        # Crear formatos con colores
        colored_formats = self._create_colored_xlsx_formats(workbook, color_mode)

        # Obtener líneas del reporte
        print_mode_self = self.with_context(no_format=True)
        lines = self._filter_out_folded_children(print_mode_self._get_lines(options))

        # Configurar columnas iniciales
        account_lines_split_names = {}
        for line in lines:
            line_model = self._get_model_info_from_id(line['id'])[0]
            if line_model == 'account.account':
                account_lines_split_names[line['id']] = self.env['account.account']._split_code_name(line['name'])

        # Configurar anchos de columna
        if len(account_lines_split_names) > 0:
            sheet.set_column(0, 0, 15)  # Código
            sheet.set_column(1, 1, 50)  # Nombre
        else:
            sheet.set_column(0, 0, 50)

        original_x_offset = 1 if len(account_lines_split_names) > 0 else 0
        y_offset = 0
        x_offset = original_x_offset + 1

        # Formato de encabezado
        title_format = workbook.add_format({
            'font_name': 'Lato', 'font_size': 12, 'bold': True,
            'bottom': 2, 'bg_color': '#f8f9fa'
        })

        # Escribir encabezados de columna
        for column_header in options.get('column_headers', []):
            for header in column_header:
                colspan = header.get('colspan', 1)
                if colspan == 1:
                    sheet.write(y_offset, x_offset, header.get('name', ''), title_format)
                else:
                    sheet.merge_range(y_offset, x_offset, y_offset, x_offset + colspan - 1,
                                     header.get('name', ''), title_format)
                x_offset += colspan
            y_offset += 1
            x_offset = original_x_offset + 1

        # Escribir encabezados de columnas de datos
        if account_lines_split_names:
            sheet.write(y_offset, x_offset - 2, "Código", title_format)
            sheet.write(y_offset, x_offset - 1, "Nombre Cuenta", title_format)

        for column in options.get('columns', []):
            colspan = column.get('colspan', 1)
            sheet.write(y_offset, x_offset, column.get('name', ''), title_format)
            x_offset += colspan

        y_offset += 1

        # Escribir líneas del reporte con colores
        for line in lines:
            level = line.get('level', 0)
            x_offset = 0

            # Determinar si es línea de total
            line_class = line.get('class', '')
            is_total = 'total' in line.get('id', '').lower() or 'total' in line_class

            # Obtener formatos según nivel
            text_format = colored_formats[f'text_level_{min(level, 6)}']
            if is_total:
                text_format = colored_formats['total_text']

            # Código de cuenta (si aplica)
            if account_lines_split_names:
                if line['id'] in account_lines_split_names:
                    code, name = account_lines_split_names[line['id']]
                    sheet.write(y_offset, x_offset, code, text_format)
                    x_offset += 1
                    sheet.write(y_offset, x_offset, name, text_format)
                else:
                    sheet.write(y_offset, x_offset, '', text_format)
                    x_offset += 1
                    sheet.write(y_offset, x_offset, line.get('name', ''), text_format)
            else:
                # Nombre con indentación
                name = line.get('name', '')
                indent = '  ' * level
                sheet.write(y_offset, x_offset, f"{indent}{name}", text_format)

            x_offset += 1

            # Columnas numéricas con colores
            for col_data in line.get('columns', []):
                value = col_data.get('no_format')
                figure_type = col_data.get('figure_type', 'monetary')

                if figure_type in ('monetary', 'float', 'integer', 'percentage') and value is not None:
                    try:
                        num_value = float(value)

                        # Seleccionar formato según valor
                        if is_total:
                            if num_value > 0:
                                fmt = colored_formats['total_positive']
                            elif num_value < 0:
                                fmt = colored_formats['total_negative']
                            else:
                                fmt = colored_formats['total_zero']
                        else:
                            level_key = min(level, 6)
                            if num_value > 0:
                                fmt = colored_formats[f'positive_level_{level_key}']
                            elif num_value < 0:
                                fmt = colored_formats[f'negative_level_{level_key}']
                            else:
                                fmt = colored_formats[f'zero_level_{level_key}']

                        # Escribir valor con o sin símbolo
                        if use_symbols and num_value != 0:
                            symbol = '▲ ' if num_value > 0 else '▼ '
                            # Con símbolos escribimos como texto formateado
                            if num_value < 0:
                                sheet.write(y_offset, x_offset, f"{symbol}({abs(num_value):,.2f})", fmt)
                            else:
                                sheet.write(y_offset, x_offset, f"{symbol}{num_value:,.2f}", fmt)
                        else:
                            # Sin símbolos, escribir número con formato
                            if num_value < 0:
                                # Para negativos con paréntesis, usar valor absoluto
                                sheet.write_number(y_offset, x_offset, abs(num_value), fmt)
                            else:
                                sheet.write_number(y_offset, x_offset, num_value, fmt)

                    except (ValueError, TypeError):
                        sheet.write(y_offset, x_offset, col_data.get('name', ''), text_format)
                else:
                    # Valor de texto
                    sheet.write(y_offset, x_offset, col_data.get('name', '') or '', text_format)

                x_offset += 1

            y_offset += 1

        # Ajustar anchos de columna
        col_count = len(options.get('columns', []))
        for i in range(2, 2 + col_count):
            sheet.set_column(i, i, 18)

    def _create_colored_xlsx_formats(self, workbook, color_mode='accounting'):
        """
        Crea los formatos de Excel con colores para reportes colombianos.

        Args:
            workbook: xlsxwriter.Workbook
            color_mode: 'accounting', 'financial', o 'simple'

        Returns:
            dict: Diccionario de formatos
        """
        # Definir colores según modo
        if color_mode == 'accounting':
            positive_color = '#198754'  # Verde Bootstrap
            negative_color = '#dc3545'  # Rojo Bootstrap
        elif color_mode == 'financial':
            positive_color = '#0d6efd'  # Azul Bootstrap (créditos)
            negative_color = '#dc3545'  # Rojo Bootstrap (débitos)
        else:  # simple
            positive_color = '#000000'  # Negro
            negative_color = '#dc3545'  # Rojo

        formats = {}

        # Formatos por nivel (0-6)
        for level in range(7):
            is_bold = level < 3
            font_size = 12 - min(level, 2)
            indent = level

            # Texto base
            formats[f'text_level_{level}'] = workbook.add_format({
                'font_name': 'Lato',
                'font_size': font_size,
                'bold': is_bold,
                'indent': indent,
            })

            # Números positivos (verde)
            formats[f'positive_level_{level}'] = workbook.add_format({
                'font_name': 'Lato',
                'font_size': font_size,
                'bold': is_bold,
                'num_format': '#,##0.00',
                'align': 'right',
                'font_color': positive_color,
            })

            # Números negativos (rojo con paréntesis)
            formats[f'negative_level_{level}'] = workbook.add_format({
                'font_name': 'Lato',
                'font_size': font_size,
                'bold': is_bold,
                'num_format': '(#,##0.00)',
                'align': 'right',
                'font_color': negative_color,
            })

            # Números cero (gris)
            formats[f'zero_level_{level}'] = workbook.add_format({
                'font_name': 'Lato',
                'font_size': font_size,
                'bold': is_bold,
                'num_format': '#,##0.00',
                'align': 'right',
                'font_color': '#999999',
            })

        # Formatos para totales
        total_bg = '#e9ecef'
        formats['total_text'] = workbook.add_format({
            'font_name': 'Lato',
            'font_size': 12,
            'bold': True,
            'bg_color': total_bg,
            'border': 1,
            'border_color': '#dee2e6',
        })

        formats['total_positive'] = workbook.add_format({
            'font_name': 'Lato',
            'font_size': 12,
            'bold': True,
            'num_format': '#,##0.00',
            'align': 'right',
            'font_color': positive_color,
            'bg_color': total_bg,
            'border': 1,
            'border_color': '#dee2e6',
        })

        formats['total_negative'] = workbook.add_format({
            'font_name': 'Lato',
            'font_size': 12,
            'bold': True,
            'num_format': '(#,##0.00)',
            'align': 'right',
            'font_color': negative_color,
            'bg_color': total_bg,
            'border': 1,
            'border_color': '#dee2e6',
        })

        formats['total_zero'] = workbook.add_format({
            'font_name': 'Lato',
            'font_size': 12,
            'bold': True,
            'num_format': '#,##0.00',
            'align': 'right',
            'font_color': '#666666',
            'bg_color': total_bg,
            'border': 1,
            'border_color': '#dee2e6',
        })

        return formats


class IrActionsReportExtend(models.Model):
    """Extiende ir.actions.report para soportar paperformat_id en contexto."""
    _inherit = 'ir.actions.report'

    def get_paperformat(self):
        """Retorna paperformat del contexto si existe, sino el por defecto."""
        paperformat_id = self._context.get('paperformat_id')
        if paperformat_id:
            return self.env['report.paperformat'].browse(paperformat_id)
        return super().get_paperformat()
